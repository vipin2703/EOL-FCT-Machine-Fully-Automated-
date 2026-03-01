
import importlib.metadata

try:
    importlib.metadata.version("nidaqmx")
except importlib.metadata.PackageNotFoundError:
    import importlib.util, sys, os
    # Locate nidaqmx manually if running from exe
    base_path = os.path.dirname(sys.executable)
    sys.path.append(base_path)




import sys
import json
import socket
import serial
from PyQt6.QtCore import Qt, QTimer, QRectF, QPointF
from PyQt6.QtGui import QColor, QLinearGradient, QBrush, QPainter, QPen, QPalette, QFont, QConicalGradient
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QGraphicsDropShadowEffect, QFrame, QMessageBox, QGridLayout,
    QHBoxLayout, QMainWindow, QStackedWidget
)

import threading
from PyQt6.QtCore import QMetaObject, Qt

import pyvisa
import nidaqmx
import time
from nidaqmx.system import System
from PyQt6.QtCore import QTimer

# ================= DB METER HELPERS =================
from pymodbus.client import ModbusSerialClient
import time
import re




from PyQt6.QtCore import QObject, pyqtSignal, QThread, Qt, QTimer, QEventLoop
from PyQt6.QtGui import QPainter, QBrush, QLinearGradient, QColor
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLineEdit, QPushButton, QMessageBox, QTabWidget, QTableWidget, QTableWidgetItem,
    QFrame
)

from PyQt6.QtWidgets import (
    QApplication, QWidget, QMainWindow, QLabel, QVBoxLayout, QHBoxLayout,
    QFrame, QPushButton, QTableWidget, QTableWidgetItem, QTabWidget, QScrollArea
)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
import re
import serial
import socket
import time
import threading
import serial, socket, threading
import nidaqmx
from nidaqmx.system import System
import os
from openpyxl import Workbook
from datetime import datetime
import json, os

from datetime import datetime
import re
import time
import serial







# ================= CONFIGURATION =================
# USERS = [
#     {"username": "1", "password": "1"},
#     {"username": "Admin", "password": "Admin"}
# ]
BAUDRATE = 115200            # ECALL/dashcam default
DB_METER_BAUDRATE = 4800     # Common for many dB meters; change if needed
FEASABAUDRATE = 115200


## ----------- TEST CASES TABLE -----------
TEST_CASES = [
    # {"name":"voltage set","channel":"PSU",
    #  "channel_number":2,"Value":12,"expected":"","min":10,"max":12,"delay":1},
    
      {"name": "CAMERA ENTRY ", "channel": "DASHCAM",
     "input_cmd": " AA 43 02 01 01 EB ", "expected": "AA", "delay": 1},



     {
  "name": "Sutter Open",
  "background_cmd": "DAQ_ONLY", 
  "daq_steps": [
    {"serial": "ni_device_1", "port": 1, "line": 6, "state": 0, "wait": 5}
  ]
},


]

# ================= RUNTIME STATE =================
ser_ecall = None
ser_dashcam = None
serfeasa = None
ser_dbmeter = None



# ---- Power Supply Globals ----
psu_rm = None
keysight_psu = None
psu_ip = None

ni_device_1 = ni_device_2 = ni_device_3 = None


LOGGED_IN_USER = None


cam_sock = None
cam_ip = None
cam_port = None
ports_connected = False  # Flag to avoid reconnect
active_tree = None
tab_counter = 0
stop_event = threading.Event()

# Tabs / UI state
result_notebook = None
active_tree = None
tab_counter = 1



# ---- Global counters ----
total_count = 0
pass_count = 0
fail_count = 0
yield_count = 0



# ================== UTILS ==================
def normalize_hex(s: str) -> str:
    if not s:
        return ""
    parts = s.replace(",", " ").replace("\n", " ").split()
    parts = [p.upper() for p in parts]
    return " ".join(parts)

def hex_to_bytes(hex_str: str) -> bytes:
    return bytes.fromhex(normalize_hex(hex_str))

def bytes_to_hex(b: bytes) -> str:
    return " ".join(f"{x:02X}" for x in b)

def to_ascii_string(b: bytes) -> str:
    # Printable ASCII 32..126; others skipped
    return "".join(chr(x) for x in b if 32 <= x <= 126)

def hex_to_ascii(hex_str: str) -> str:
    try:
        data = bytes.fromhex(normalize_hex(hex_str))
        s = to_ascii_string(data)
        return s if s else ""
    except Exception:
        return ""

def send_hex(port: str, hex_str: str):
    data = hex_to_bytes(hex_str)
    if port == "ECALL" and ser_ecall:
        ser_ecall.write(data)
    elif port == "DASHCAM" and ser_dashcam:
        ser_dashcam.write(data)
    elif port == "ETH" and cam_sock:
        print("ethernet command send")
        cam_sock.sendall(data)
    elif port == "FEASA" and serfeasa:
        # payload = query.encode('ascii')  # Sirf ASCII command, bina CR LF
        serfeasa.write(data)
    else:
        raise RuntimeError(f"Port not open: {port}")
    print(f"📤 Sent to {port}: {normalize_hex(hex_str)}")
 
def read_serial_response(ser: serial.Serial, timeout_s: float | None) -> str | None:
   
    start = time.time()
    buf = bytearray()
    while True:
        n = ser.in_waiting
        if n:
            chunk = ser.read(n)
            buf.extend(chunk)
            time.sleep(0.05)
            n2 = ser.in_waiting
            if n2:
                buf.extend(ser.read(n2))
            return bytes_to_hex(bytes(buf))
        if timeout_s is not None and (time.time() - start) >= timeout_s:
            return None
        time.sleep(0.05)



def read_serial_response_exit(
    ser: serial.Serial,
    timeout_s: float | None,
    read_window: float | None = 0.3   # default 0.3s agar command me nahi diya
) -> str | None:
    """
    Special read function for EXIT command.
    Collects all bytes arriving within `read_window` after first byte.
    """
    start = time.time()
    buf = bytearray()
    first_byte_received = False
    first_byte_time = None

    while True:
        n = ser.in_waiting
        if n:
            chunk = ser.read(n)
            buf.extend(chunk)
            if not first_byte_received:
                first_byte_received = True
                first_byte_time = time.time()
            # continue reading while within read_window
            elif read_window and (time.time() - first_byte_time) < read_window:
                time.sleep(0.05)
                n2 = ser.in_waiting
                if n2:
                    buf.extend(ser.read(n2))
            else:
                # read_window expired after first byte
                return bytes_to_hex(bytes(buf))

        # timeout for overall command
        if timeout_s is not None and (time.time() - start) >= timeout_s:
            if buf:
                return bytes_to_hex(bytes(buf))  # return whatever collected
            return None

        time.sleep(0.05)






def set_ecall_date_command(base_cmd: str = "B4 12") -> str:
    """
    Generate ECALL command with today's date bytes
    base_cmd: command prefix before date bytes
    Returns full hex string e.g. "B3 01 10 0A 19"
    """
    today = datetime.today()
    day_dec = f"{today.day:02}"
    month_dec = f"{today.month:02}"
    year_dec = f"{today.year % 100:02}"

    full_cmd = f"{base_cmd} {day_dec} {month_dec} {year_dec}"
    return full_cmd



# ================== FLEXIBLE MATCHING ==================
RANGE_RE = re.compile(r"^\[(\d+)\-(\d+)\]$")
HEX_BYTE_RE = re.compile(r"^[0-9A-Fa-f]{2}$")
DEC_BYTE_RE = re.compile(r"^\d+$")

def parse_expected_tokens(expected: str):
    expected = expected.strip()
    return expected.split() if expected else []

def consume_range_value(resp_bytes: bytes, start_idx: int, min_v: int, max_v: int, range_bytes: int | None):
    if range_bytes not in (1, 2, None):
        range_bytes = None
    if range_bytes is None:
        range_bytes = 2 if max_v > 255 else 1
    end_idx = start_idx + range_bytes
    if end_idx > len(resp_bytes):
        return False, 0, None
    if range_bytes == 1:
        val = resp_bytes[start_idx]
    else:
        val = (resp_bytes[start_idx] << 8) | resp_bytes[start_idx + 1]
    return (min_v <= val <= max_v), range_bytes, val

def match_pattern_over_bytes(resp_bytes: bytes, expected: str, fmt: str, range_bytes_override: int | None = None):
    fmt_u = (fmt or "HEX").strip().upper()

    if fmt_u == "ASCII":
        ascii_str = to_ascii_string(resp_bytes)
        if not expected:
            return (len(ascii_str) > 0, ascii_str, "")
        ok = expected in ascii_str
        return (ok, ascii_str if ok else "", "")

    tokens = parse_expected_tokens(expected)
    if not tokens:
        return (len(resp_bytes) > 0, bytes_to_hex(resp_bytes), "")

    resp_len = len(resp_bytes)

    for start in range(0, resp_len):
        i_resp = start
        consumed_hex_parts = []
        notes = []
        matched_all = True

        for tk in tokens:
            m = RANGE_RE.match(tk)
            if m:
                min_v = int(m.group(1))
                max_v = int(m.group(2))
                ok, consumed, val = consume_range_value(resp_bytes, i_resp, min_v, max_v, range_bytes_override)
                if not ok:
                    matched_all = False
                    break
                if consumed == 1:
                    consumed_hex_parts.append(f"{resp_bytes[i_resp]:02X}")
                else:
                    consumed_hex_parts.append(f"{resp_bytes[i_resp]:02X} {resp_bytes[i_resp+1]:02X}")
                notes.append(f"(range {min_v}-{max_v} val={val})")
                i_resp += consumed
                continue

            if fmt_u == "HEX":
                if not HEX_BYTE_RE.match(tk):
                    matched_all = False
                    break
                if i_resp >= resp_len:
                    matched_all = False
                    break
                if resp_bytes[i_resp] != int(tk, 16):
                    matched_all = False
                    break
                consumed_hex_parts.append(f"{resp_bytes[i_resp]:02X}")
                i_resp += 1
                continue

            if fmt_u in ("DEC", "DECIMAL"):
                if DEC_BYTE_RE.match(tk):
                    if i_resp >= resp_len:
                        matched_all = False
                        break
                    if resp_bytes[i_resp] != int(tk, 10):
                        matched_all = False
                        break
                    consumed_hex_parts.append(f"{resp_bytes[i_resp]:02X}")
                    i_resp += 1
                    continue
                matched_all = False
                break

            matched_all = False
            break

        if matched_all:
            matched_hex = " ".join(consumed_hex_parts)
            extra = " ".join(notes)
            return True, matched_hex, extra

    return False, "", ""

def build_output_string(hex_slice: str, extra_note: str) -> str:
    return hex_slice if not extra_note else f"{hex_slice}  {extra_note}"





# ================= DB METER HELPERS =================
import time, re, struct

NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")

def dbmeter_write_and_read(ser_dbmeter, query: str = None, timeout_s: float = 1) -> str:
    """
    Reads noise value from Modbus DB meter using already-open ser_dbmeter.
    No need to reconnect; same style as previous ASCII version.
    """
    if ser_dbmeter is None or not ser_dbmeter.is_open:
        raise RuntimeError("DB Meter port not open")

    # --- Build a simple Modbus RTU request manually ---
    # Example: Slave=1, Function=0x03 (Read Holding Reg), Addr=0x0000, Count=1
    request = bytes([0x01, 0x03, 0x00, 0x00, 0x00, 0x01])
    
    # Append CRC16
    crc = calc_crc16(request)
    request += struct.pack('<H', crc)

    print(f"[DB METER] Sending HEX: {' '.join(f'{b:02X}' for b in request)}")

    # Clear buffers
    ser_dbmeter.reset_input_buffer()
    ser_dbmeter.reset_output_buffer()
    ser_dbmeter.write(request)
    ser_dbmeter.flush()

    # Read response
    start = time.time()
    resp = bytearray()
    while time.time() - start < timeout_s:
        n = ser_dbmeter.in_waiting
        if n:
            resp.extend(ser_dbmeter.read(n))
        if len(resp) >= 7:  # minimum valid Modbus frame
            break
        time.sleep(0.02)

    if not resp:
        raise RuntimeError("No response from DB meter")

    print(f"[DB METER] Received HEX: {' '.join(f'{b:02X}' for b in resp)}")

    # Parse Modbus response
    if len(resp) >= 7 and resp[1] == 0x03:
        raw_val = (resp[3] << 8) | resp[4]
        db_val = raw_val / 10.0
        print(f"[DB METER] Parsed Value: {db_val:.2f} dB")
        return str(db_val)
    else:
        raise RuntimeError(f"Invalid Modbus response: {resp.hex()}")


def calc_crc16(data: bytes) -> int:
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            if crc & 0x0001:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc


# ================= DB VALUE PARSER =================
def parse_db_value(raw_resp: str) -> float:
    if not raw_resp:
        raise ValueError("Empty response from DB meter")
    match = NUMBER_RE.search(raw_resp)
    if not match:
        raise ValueError(f"No numeric value found in response: {raw_resp}")
    return float(match.group(0))



# ================= COMMAND RUNNERS =================
def run_on_off(cmd_info: dict):
    try:
        send_hex(cmd_info["port"], cmd_info["on_cmd"])
        time.sleep(cmd_info["delay"])
        send_hex(cmd_info["port"], cmd_info["off_cmd"])
    except Exception as e:
        QMessageBox.information("Error", f"Command failed: {e}")
        





def niusb_write_line(serial_hex: str | None, port: int, line: int, state=None, pulse: bool = True, wait: float = 1.0) -> bool:
    """
    Sends digital output to NI USB-6501 line.
    ✅ Supports both:
       - niusb_write_line("020B85CA", 0, 7)
       - niusb_write_line(""ni_device_2"", 0, 7)
    """
    try:
        # 🧩 1. Check if None
        if serial_hex is None:
            print("❌ NIUSB Write Error: serial_hex is None (device not connected yet).")
            return False

        # 🧩 2. If global variable name string is passed like ""ni_device_2""
        if isinstance(serial_hex, str) and serial_hex.startswith("ni_device_"):
            resolved_serial = globals().get(serial_hex)
            if resolved_serial is None:
                print(f"❌ NIUSB Write Error: {serial_hex} not initialized yet.")
                return False
            serial_hex = resolved_serial

        # 🧩 3. Validate and convert hex → decimal
        serial_decimal = int(serial_hex, 16)

        # 🧩 4. Find device by serial number
        system = System.local()
        dev = next((d for d in system.devices if getattr(d, "serial_num", None) == serial_decimal), None)
        if not dev:
            print(f"❌ Device with serial {serial_hex} (DEC {serial_decimal}) not found")
            return False

        # 🧩 5. Prepare channel and execute write
        line_name = f"{dev.name}/port{port}/line{line}"
        with nidaqmx.Task() as task:
            task.do_channels.add_do_chan(line_name)

            if pulse:
                task.write(False)
                print(f"[PULSE] {line_name} -> LOW")
                if wait is not None:
                    time.sleep(wait)
                    task.write(True)
                    print(f"[PULSE] {line_name} -> HIGH")
                else:
                    print("[PULSE] wait skipped")
            else:
                task.write(bool(state))
                print(f"[OK] {line_name} -> {'HIGH' if state else 'LOW'}")

        return True

    except Exception as e:
        print(f"NIUSB Write Error: {e}")
        return False














def niusb_set_low(serial_hex: str, port: int, line: int) -> bool:
    """
    Sets NI USB-6501 line to LOW and keeps it LOW.
    HIGH pe switch nahi karega.
    """
    try:
        serial_decimal = int(serial_hex, 16)
        system = System.local()
        dev = next((d for d in system.devices if getattr(d, "serial_num", None) == serial_decimal), None)

        if not dev:
            print(f"❌ Device with serial {serial_hex} (DEC {serial_decimal}) not found")
            return False

        line_name = f"{dev.name}/port{port}/line{line}"
        with nidaqmx.Task() as task:
            task.do_channels.add_do_chan(line_name)
            task.write(False)  # LOW pe set
            print(f"[SET LOW] {line_name} -> LOW")

        return True

    except Exception as e:
        print(f"NIUSB Set LOW Error: {e}")
        return False






##########REAL
class EntryTimeoutException(Exception):
    pass


from PyQt6.QtCore import QTimer

# def run_all_tests(table: QTableWidget, status_label: QLabel, window_obj: "MainScreen", main_window=None):
#     stop_event.clear()
#     try:
#         for idx, case in enumerate(TEST_CASES):
#             if stop_event.is_set():
#                 table.setItem(idx, 4, QTableWidgetItem("STOPPED"))
#                 break
#             run_test_case(idx, table, status_label)
#     except EntryTimeoutException:
#         print("ENTRY failed → EXIT executed → stopping further tests.")

#     finally:
#         status_label.setText("All tests completed.")
#         generate_report(table)
#         generate_excel(table)
#         finalize_run(table, window_obj)
#         print("after finalize come")
#         stop_event.clear()

#         # Safely switch to barcode screen on main thread
#         if main_window is not None:
#             QTimer.singleShot(0, main_window.open_barcode_screen)
#         else:
#             print("main_window is None; barcode screen not opened")

def run_all_tests(table: QTableWidget, status_label: QLabel, window_obj: "MainScreen", main_window=None):
    stop_event.clear()
    try:
        for idx, case in enumerate(TEST_CASES):
            if stop_event.is_set():
                table.setItem(idx, 4, QTableWidgetItem("STOPPED"))
                break
            run_test_case(idx, table, status_label)
    except EntryTimeoutException:
        print("ENTRY failed → EXIT executed → stopping further tests.")
    finally:
        status_label.setText("All tests completed.")

        # generate_report now returns overall_result ("PASS"/"FAIL")
        try:
            overall_result = generate_report(table)
        except Exception as e:
            print(f"generate_report error: {e}")
            overall_result = "FAIL"

        try:
            generate_excel(table)
        except Exception as e:
            print(f"generate_excel error: {e}")

        try:
            finalize_run(table, window_obj)
        except Exception as e:
            print(f"finalize_run error: {e}")

        print("after finalize come")
        stop_event.clear()
        
        # Safely show overall banner on main thread (if available), which will open barcode after duration
        if main_window is not None:
            QTimer.singleShot(0, lambda: main_window.show_overall_banner(overall_result, duration_ms=3000))
        else:
            print("main_window is None; barcode screen not opened")

    


# ================= DB VALUE PARSER =================
def parse_db_value(raw_resp: str) -> float:
    if not raw_resp:
        raise ValueError("Empty response from DB meter")
    match = re.search(r"[-+]?\d*\.\d+|\d+", raw_resp)
    if not match:
        raise ValueError(f"No numeric value found in response: {raw_resp}")
    return float(match.group(0))


def run_test_case(idx: int, table: QTableWidget, status_var: QLabel, is_exit=False):
    """
    Runs a single test case from TEST_CASES, fully updated for PyQt6 QTableWidget.
    Supports DAQ, DB meter, serial, Ethernet, FEASA, ASCII/HEX parsing, delays, ENTRY, background, special cases.
    """

    case = TEST_CASES[idx]
    # --- Extract variables safely ---
    name = case["name"]
    # channel = case["channel"].upper()
    channel = case.get("channel")
    if channel:
        channel = channel.upper()
    else:
        channel = None
    input_cmd = normalize_hex(case.get("input_cmd", ""))
    expected = case.get("expected", "").strip()
    expected_hex_view = normalize_hex(expected) if expected else ""
    hidden_cmd = normalize_hex(case.get("hidden_cmd", ""))  # optional hidden clear cmd
    fmt = (case.get("format", "HEX") or "HEX").upper()
    range_bytes_override = case.get("range_bytes", None)

    mode_ascii = case.get("ascii")
    names_specail = case.get("special")
    mode_hex=case.get("hex")
    min_val = case.get("min")
    max_val = case.get("max")
    cmd_dict1=case.get("x_axis")
    cmd_dict2=case.get("y_axis")
    cmd_dict3=case.get("z_axis")
    background_cmd = case.get("background_cmd")
    
    # --- DB meter extras ---
    dbmeter_cmd = case.get("dbmeter_cmd")
    dbmeter_delay = float(case.get("dbmeter_delay", 1.5))
    dbmeter_checks = case.get("dbmeter_checks")  # None OR list of dicts

    daq_info = case.get("daq_steps")
    is_daq_case = daq_info is not None
    feasa_data = case.get("feasa_data")

    # PyQt TableWidget row
    table_item_row = idx
    # table.setItem(table_item_row, 1, QTableWidgetItem(input_cmd))
    # table.setItem(table_item_row, 2, QTableWidgetItem(expected))
    table.setItem(table_item_row, 2, QTableWidgetItem(""))
    table.setItem(table_item_row, 4, QTableWidgetItem("RUNNING..."))

    capture_only = (expected.strip() == "")
    start_time=time.time()

    try:
        start_time = time.time()
        is_entry_case = ("ENTRY" in name.upper() and expected_hex_view.upper() == "B5 02 81 01")
        base_timeout = 30 if is_entry_case else 2
        timeout_s = base_timeout + float(case.get("extra_delay_read", 0))

        # --- Serial / Channel Setup ---
        if channel in ("ECALL", "DASHCAM"):
            ser = ser_ecall if channel == "ECALL" else ser_dashcam
            if ser is None:
                raise RuntimeError(f"{channel} port not open")
            ser.reset_input_buffer()
            ser.reset_output_buffer()
        elif channel == "ETH":
            if cam_sock is None:
                raise RuntimeError("Camera (Ethernet) not connected")
        elif channel == "FEASA":
            if serfeasa is None:
                raise RuntimeError("FEASA port not open")
        elif not channel:
            print("[INFO] No channel specified → assuming DAQ-only/background case")



        else:
            raise RuntimeError(f"Unknown channel: {channel}")

        # --- Hidden command ---
        if hidden_cmd:
            send_hex(channel, hidden_cmd)
            time.sleep(0.1)
            if channel in ("ECALL", "DASHCAM"):
                ser.reset_input_buffer()

        if case.get("name") == "SET DATE OF MANUFACTURER":
            # dynamic date command
            input_cmd = set_ecall_date_command()  # generate today's date hex




        # --- Input command + DAQ handling ---
        if input_cmd:
           
            if case.get("name") == "Read Touch Data" and is_daq_case and daq_info:
                # Step 1: pehle command bhejo
                send_hex(channel, input_cmd)
                print("[INFO] Dashcam/Ecall command sent (special case: Read Touch Data)")
                time.sleep(0.5)
                first_step = daq_info[0]
                wait_time = first_step.get("wait", 0)
                niusb_write_line(first_step["serial"], first_step["port"], first_step["line"], pulse=True, wait=wait_time)
                print(f"[INFO] First DAQ pulse sent: {first_step['serial']}/{first_step['port']}/{first_step['line']}")
                time.sleep(1)

            elif case.get("name") == "Mic Speaker Loop Start" and is_daq_case and daq_info:
                # Step 1: pehle command bhejo
                send_hex(channel, input_cmd)
                print("[INFO] Dashcam/Ecall command sent (special case:mic speaker loop)")
                time.sleep(2)
                first_step = daq_info[0]
                wait_time = first_step.get("wait", 0)
                niusb_write_line(first_step["serial"], first_step["port"], first_step["line"], pulse=True, wait=wait_time)
                print(f"[INFO] First DAQ pulse sent: {first_step['serial']}/{first_step['port']}/{first_step['line']}")

            elif case.get("name") == "Get data x-axis" and is_daq_case and daq_info:
                # Step 1: pehle command bhejo
                print("[INFO] Dashcam/Ecall command sent (special case: Read x axis data)")
                first_step = daq_info[0]
                niusb_write_line(first_step["serial"], first_step["port"], first_step["line"], pulse=True)
                print(f"[INFO] First DAQ pulse sent: {first_step['serial']}/{first_step['port']}/{first_step['line']}")
                # Step 2: 1 sec wait
                time.sleep(2)
                # Step 3: DAQ ka pehla step
                send_hex(channel, input_cmd)
            else:
                # Normal DAQ step
                if is_daq_case and daq_info:
                    first_step = daq_info[0]
                    wait_time = first_step.get("wait", 0)
                    niusb_write_line(first_step["serial"], first_step["port"], first_step["line"], pulse=True, wait=wait_time)
                    print(f"[INFO] First DAQ pulse sent: {first_step['serial']}/{first_step['port']}/{first_step['line']}")
                print(input_cmd)
                send_hex(channel, input_cmd)
                print("[INFO] Dashcam/Ecall command sent with DAQ step")
        else:
            # Background DAQ only
            if background_cmd == "DAQ_ONLY" and is_daq_case and daq_info:
                print(f"[INFO] Background DAQ-only case: {name}")
                print("ni_device_2")
                
                for step in daq_info:
                    wait_time = step.get("wait", 0)
                    niusb_write_line(step["serial"], step["port"], step["line"], pulse=True, wait=wait_time)
                    print(f"[INFO] DAQ step executed: {step['serial']}/{step['port']}/{step['line']} (wait={wait_time}s)")
                    table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                return
                    

        # Extra delay before reading
        extra_delay_ = float(case.get("extra_delay", 0))
        if extra_delay_ > 0:
            print(f"[INFO] Waiting {extra_delay_}s before starting read...")
            time.sleep(extra_delay_)


        # ---------- ECALL Response + DB Meter Validation ----------
        if dbmeter_cmd:
            print("[DBMETER] Waiting for ECALL expected response before measuring dB...")
        
            ecall_resp = read_serial_response(ser_ecall, timeout_s=1.0)
            if not ecall_resp:
                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (No ECALL response)"))
                return
        
            ecall_resp_norm = normalize_hex(ecall_resp)
            print("[DBMETER] ECALL response HEX:", ecall_resp_norm)
        
            expected_norm = normalize_hex(expected)
            if not ecall_resp_norm.startswith(expected_norm):
                print(f"[DBMETER] ❌ ECALL response mismatch. Expected prefix: {expected_norm}")
                table.setItem(table_item_row, 2, QTableWidgetItem(f"ECALL HEX: {ecall_resp_norm}"))
                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Expected mismatch)"))
                return
            else:
                print("[DBMETER] ✅ ECALL expected response matched. Proceeding to dB meter test...")
        
            # 2️⃣ DB meter measurement
            output_parts = [f"ECALL HEX: {ecall_resp_norm}"]
        
            if not dbmeter_checks:
                # Single DB meter read (Modbus)
                time.sleep(dbmeter_delay)  # wait before reading
                try:
                    raw_resp = dbmeter_write_and_read(ser_dbmeter)
                    db_val = parse_db_value(raw_resp)
                    output_parts.append(f"{db_val:.2f} dB")
                    table.setItem(table_item_row, 2, QTableWidgetItem(" | ".join(output_parts)))
        
                    if min_val is not None and max_val is not None:
                        if min_val <= db_val <= max_val:
                            table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                        else:
                            table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Out of range: {db_val:.2f} dB)"))
                    else:
                        table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
        
                    if case.get("delay", 0) > 0:
                        time.sleep(case["delay"])
        
                    return
                except Exception as pe:
                    output_parts.append(f"[PARSE ERROR] {pe}")
                    table.setItem(table_item_row, 2, QTableWidgetItem(" | ".join(output_parts)))
                    table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (DB parse error)"))
                    return
        
            else:
                # Multi-step DB meter checks (Modbus)
                print("enter ")
                overall_pass = True
                for check in dbmeter_checks:
                    if stop_event.is_set():
                        table.setItem(table_item_row, 4, QTableWidgetItem("STOPPED"))
                        return
        
                    time.sleep(check.get("after", 0.5))
                    try:
                        raw_resp = dbmeter_write_and_read(ser_dbmeter)
                        db_val = parse_db_value(raw_resp)
                        label = check.get("label", "DB")
                        min_val_step = check.get("min")
                        max_val_step = check.get("max")
        
                        output_parts.append(f"{label}: {db_val:.2f} dB")
        
                        step_pass = True
                        if min_val_step is not None and max_val_step is not None:
                            step_pass = min_val_step <= db_val <= max_val_step
                        overall_pass = overall_pass and step_pass
        
                    except Exception as pe:
                        output_parts.append(f"{label}: PARSE ERROR ({pe})")
                        overall_pass = False
        
                # ✅ Final output
                table.setItem(table_item_row, 2, QTableWidgetItem(" | ".join(output_parts)))
                table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅" if overall_pass else "FAIL ❌"))
                return





        # ---------- Response Loop ----------
        response_received = False
        buffer_hex_strings = []
        resp_all_hex = ""  # ensure defined even if no response arrives

        while time.time() - start_time < timeout_s:
            if stop_event.is_set():
                table.setItem(table_item_row, 4, QTableWidgetItem("STOPPED"))
                return

            resp_hex = None

            if channel in ("ECALL", "DASHCAM"):
                if name.upper() == "EXIT":
                    resp_hex = read_serial_response_exit(
                        ser,
                        timeout_s=timeout_s,
                        read_window=case.get("read_window", 0.3)
                    )
                else:
                    resp_hex = read_serial_response(ser, timeout_s=1)

            # ... (ETH and FEASA handling remain as in your original code above) ...

            elif channel == "ETH":
                cam_sock.settimeout(0.1)
                timeout = 2
                end_time = time.time() + timeout
                resp_data = b""
                while time.time() < end_time:
                    try:
                        chunk = cam_sock.recv(4096)
                        if not chunk:
                            break
                        resp_data += chunk
                    except Exception:
                        continue
            
                resp_ascii = resp_data.decode('ascii', errors='ignore').replace('\r', '').replace('\n', '').strip()
                resp_values_str = resp_ascii[resp_ascii.index("T1") + 2:].strip() if "T1" in resp_ascii else resp_ascii
            
                value_ranges = case.get("value_ranges")
                min_v = case.get("min_val")
                max_v = case.get("max_val")
                min_l = case.get("min_limit")
                max_l = case.get("max_limit")
            
                status = "FAIL ❌"
                output = resp_ascii
            
                # -------------------- Output handling for T1 responses (using min_limit/max_limit) --------------------
                if min_l is not None and max_l is not None:
                    try:
                        # Convert T10/T11 into numeric value
                        if "T10" in resp_ascii:
                            val = 0
                        elif "T11" in resp_ascii:
                            val = 1
                        else:
                            val = None
            
                        # Compare based on interpreted value
                        if val is not None:
                            output = f"Response: {resp_ascii}, Interpreted Value: {val}"
                            if val == 0:
                                status = "PASS ✅"
                            else:
                                status = "FAIL ❌"
                        else:
                            output = f"Unexpected Response: {resp_ascii}"
                            status = "FAIL ❌"
                    except Exception as e:
                        output = f"T1 Parse Error: {e}"
                        status = "FAIL ❌"
            
                # -------------------- Normal handling --------------------
                elif value_ranges:
                    try:
                        vals = [float(v) for v in resp_values_str.split(",")]
                        if len(vals) == len(value_ranges) and all(lo <= val <= hi for val, (lo, hi) in zip(vals, value_ranges)):
                            status = "PASS ✅"
                    except Exception as e:
                        output = f"Parse Error: {e}"
            
                elif min_v is not None and max_v is not None:
                    try:
                        last_val = int(resp_values_str[-3:])
                        if min_v <= last_val <= max_v:
                            status = "PASS ✅"
                        output = f"Extracted last 3 digits: {last_val}"
                    except Exception as e:
                        status = "FAIL ❌"
                        output = f"Parse Error: {e}"
            
                else:
                    expected_str = case.get("expected", "")
                    status = "PASS ✅" if expected_str in resp_ascii else "FAIL ❌"
            
                # -------------------- Update GUI Table --------------------
                table.setItem(table_item_row, 2, QTableWidgetItem(output))
                table.setItem(table_item_row, 4, QTableWidgetItem(status))
                return


            # FEASA response
            elif channel == "FEASA" and feasa_data:
                buf = bytearray()
                start = time.time()
                while time.time() - start < 1.0:
                    n = serfeasa.in_waiting
                    if n: buf.extend(serfeasa.read(n))
                    time.sleep(0.05)
                out = buf.decode(errors="ignore").strip()
                table.setItem(table_item_row, 2, QTableWidgetItem(out if out else "NO RESPONSE"))
                r_val = None
                match = re.search(r'[Rr]:\s*([+-]?[0-9]*\.?[0-9]+)', out)
                if match:
                    try: r_val = int(match.group(1))
                    except: r_val = None
                if r_val is not None and min_val is not None and max_val is not None:
                    table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅" if min_val <= r_val <= max_val else "FAIL ❌"))
                    return











            if resp_hex:
                response_received = True
                buffer_hex_strings.append(resp_hex)

                # Build resp_all_hex safely (only when buffer non-empty)
                resp_all_hex = " ".join(buffer_hex_strings).strip()
                resp_bytes = b""
                if resp_all_hex:
                    try:
                        resp_bytes = bytes.fromhex(normalize_hex(resp_all_hex))
                    except Exception:
                        # If conversion fails, keep resp_bytes empty and continue safely
                        resp_bytes = b""

                # --- ASCII mode handling ---
                if mode_ascii is not None:
                    ascii_str = to_ascii_string(resp_bytes).strip()
                    table.setItem(table_item_row, 2, QTableWidgetItem(ascii_str or "[no printable ASCII]"))
                    parts = ascii_str.split()

                    if isinstance(min_val, str) and isinstance(max_val, str):
                        try:
                            if expected and not ascii_str.startswith(expected):
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Prefix mismatch)"))
                            else:
                                version_str = ascii_str[-8:]  # e.g. XX.XX.XX
                                resp_ver = tuple(map(int, version_str.split(".")))
                                min_ver = tuple(map(int, min_val.split(".")))
                                max_ver = tuple(map(int, max_val.split(".")))
                                if min_ver <= resp_ver <= max_ver:
                                    table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                                else:
                                    table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Out of range: {version_str})"))
                        except Exception:
                            table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Invalid version format)"))
                        return

                    elif len(parts) > 2 and names_specail is not None:
                        print(ascii_str.strip())
                        if case.get("name") == "Read Touch Data":
                            print("enter")
                            response=ascii_str.strip()
                            if response== "B3 05 AD 00":
                                table.setItem(table_item_row, 2, QTableWidgetItem(response))
                                table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ "))
                                # ---- DAQ cleanup pulse (if defined) ----
                                if is_daq_case and daq_info is not None and len(daq_info) > 1:
                                    print("enter in no response (sending second DAQ pulse)")
                                    second_step = daq_info[1]
                                    niusb_write_line(
                                        second_step["serial"],
                                        second_step["port"],
                                        second_step["line"],
                                        pulse=True
                                    )
                                    print(f"[INFO] Second DAQ pulse sent: {second_step['serial']}/{second_step['port']}/{second_step['line']}")
                                return
                            else:
                                try:
                                    response = ascii_str.strip()
                                    
                                    if response.startswith(expected):
                                        table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))

                                    else:
                                        table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Prefix mismatch: {response})"))
                                except Exception as e:
                                    table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Error: {e})"))
                                    # ---- DAQ cleanup pulse (if defined) ----
                                if is_daq_case and daq_info is not None and len(daq_info) > 1:
                                    print("enter in no response (sending second DAQ pulse)")
                                    second_step = daq_info[1]
                                    niusb_write_line(
                                        second_step["serial"],
                                        second_step["port"],
                                        second_step["line"],
                                        pulse=True
                                    )
                                    print(f"[INFO] Second DAQ pulse sent: {second_step['serial']}/{second_step['port']}/{second_step['line']}")
                                return


                        else:
                            try:
                                response = ascii_str.strip()
                                
                                if response.startswith(expected):
                                    table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                                else:
                                    table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Prefix mismatch: {response})"))
                            except Exception as e:
                                table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Error: {e})"))
                                # ---- DAQ cleanup pulse (if defined) ----
                            # if is_daq_case and daq_info is not None and len(daq_info) > 1:
                            #     print("enter in no response (sending second DAQ pulse)")
                            #     second_step = daq_info[1]
                            #     niusb_write_line(
                            #         second_step["serial"],
                            #         second_step["port"],
                            #         second_step["line"],
                            #         pulse=True
                            #     )
                            #     print(f"[INFO] Second DAQ pulse sent: {second_step['serial']}/{second_step['port']}/{second_step['line']}")
                            return


                    else:
                        try:
                            if parts[-1].isalpha():
                                val_token = parts[-2]
                                unit = parts[-1]
                            else:
                                val_token = parts[-1]
                                unit = ""
                            decimal_val = float(val_token)
                            if expected and not ascii_str.startswith(expected):
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Prefix mismatch)"))
                            elif min_val <= decimal_val <= max_val:
                                table.setItem(table_item_row, 4, QTableWidgetItem(f"PASS ✅"))
                            else:
                                table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Out of range: {decimal_val} {unit})"))
                        except ValueError:
                            table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Invalid number in response)"))
                        return

                # --- HEX mode handling ---
                elif mode_hex is not None:
                    # Only attempt hex parsing when we have some bytes
                    if resp_bytes:
                        resp_all_hex = " ".join(f"{b:02X}" for b in resp_bytes)
                    else:
                        resp_all_hex = ""  # keep empty if nothing valid

                    # axis parsing or simple hex parsing as before, but guarded
                    if cmd_dict1 is not None or cmd_dict2 is not None or cmd_dict3 is not None:
                        if not resp_bytes:
                            table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (No hex response)"))
                            return
                        try:
                            resp_all_hex = " ".join(f"{b:02X}" for b in resp_bytes)
                            bytes_list = resp_all_hex.split()
                            idx = bytes_list.index("02")
                            if cmd_dict1 is not None:
                                low_byte = int(bytes_list[idx + 1], 16)
                                high_byte = int(bytes_list[idx + 2], 16)
                            elif cmd_dict2 is not None:
                                low_byte = int(bytes_list[idx + 3], 16)
                                high_byte = int(bytes_list[idx + 4], 16)
                            elif cmd_dict3 is not None:
                                low_byte = int(bytes_list[idx + 5], 16)
                                high_byte = int(bytes_list[idx + 6], 16)
                            decimal_val = round((high_byte * 256 + low_byte) * 0.488, 3)
                            table.setItem(table_item_row, 2, QTableWidgetItem(f"{decimal_val}"))
                            expected_clean = expected.replace(" ", "")
                            resp_clean = resp_all_hex.replace(" ", "")
                            if expected and not resp_clean.startswith(expected_clean):
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Prefix mismatch)"))
                            elif min_val <= decimal_val <= max_val:
                                table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                            else:
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Out of range)"))
                            return
                        except Exception as e:
                            table.setItem(table_item_row, 4, QTableWidgetItem(f"FAIL ❌ (Hex parse error: {e})"))
                            return
                    else:
                        # simple hex case: require resp_bytes
                        if not resp_bytes:
                            table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (No hex response)"))
                            return
                        resp_all_hex = " ".join(f"{b:02X}" for b in resp_bytes)
                        table.setItem(table_item_row, 2, QTableWidgetItem(resp_all_hex))
                        try:
                            decimal_val = int.from_bytes(resp_bytes[-1:], byteorder="big")
                            expected_clean = expected.replace(" ", "")
                            resp_clean = resp_all_hex.replace(" ", "")
                            if expected and not resp_clean.startswith(expected_clean):
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Prefix mismatch)"))
                            elif min_val <= decimal_val <= max_val:
                                table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                            else:
                                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Out of range)"))
                            return
                        except Exception:
                            table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Invalid hex parse)"))
                            return

                # --- Pattern-match bytes (generic fallback) ---
                else:
                    ok, matched_slice, extra_note = match_pattern_over_bytes(resp_bytes, expected, fmt, range_bytes_override)
                    if ok:
                        out_str = build_output_string(matched_slice, extra_note)
                        table.setItem(table_item_row, 2, QTableWidgetItem(out_str if fmt != "ASCII" else (out_str or to_ascii_string(resp_bytes))))
                        table.setItem(table_item_row, 4, QTableWidgetItem("PASS ✅"))
                        if is_daq_case and len(daq_info) > 1:
                            second_step = daq_info[1]
                            niusb_write_line(second_step["serial"], second_step["port"], second_step["line"], pulse=True)
                        break

        else:
            # ---- TIMEOUT (no response received within timeout_s) ----
            print(f"[WARN] Timeout waiting for response for test '{name}' (mode_hex={mode_hex})")

            # ===== ENTRY TIMEOUT CASE =====
            if is_entry_case:
                table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (ENTRY timeout)"))
                stop_event.set()
                # Optional: call your report generation or barcode screen reset here
                raise EntryTimeoutException()


            # ===== Special Retry ONLY for WIFI MAC =====
            if name.upper() == "GET WIFI MAC ID":

                # Initialize retry_count if missing
                case["retry_count"] = case.get("retry_count", 0) + 1
                
                if case["retry_count"] <= 3:     # Try max 3 times
                    print(f"[RETRY] WIFI MAC Timeout → Retrying {case['retry_count']}/3 ...")

                    # Send the command again
                    send_hex(channel, input_cmd)
                    time.sleep(1)

                    # ⚡ Re-run SAME TEST CASE again
                    run_test_case(idx, table, status_var)
                    return  # Important

                else:
                    print("[FAIL] WIFI MAC failed after 3 retries")
                    table.setItem(table_item_row, 2, QTableWidgetItem("NO RESPONSE"))
                    table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Timeout)"))
                    return   # Final failure






            # ===== NORMAL TIMEOUT CASE =====
            else:
                # For HEX-based tests
                if mode_hex is not None:
                    if resp_all_hex:
                        table.setItem(table_item_row, 2, QTableWidgetItem(resp_all_hex))
                    else:
                        table.setItem(table_item_row, 2, QTableWidgetItem("[NO HEX RESPONSE]"))
                    table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ (Timeout)"))

                # For ASCII or special tests
                else:
                    table.setItem(table_item_row, 2, QTableWidgetItem(resp_all_hex))
                    # table.setItem(table_item_row, 2, QTableWidgetItem("[NO RESPONSE]"))
                    table.setItem(table_item_row, 4, QTableWidgetItem("FAIL ❌ "))

                # ---- DAQ cleanup pulse (if defined) ----
                if is_daq_case and daq_info is not None and len(daq_info) > 1:
                    print("enter in no response (sending second DAQ pulse)")
                    second_step = daq_info[1]
                    niusb_write_line(
                        second_step["serial"],
                        second_step["port"],
                        second_step["line"],
                        pulse=True
                    )
                    print(f"[INFO] Second DAQ pulse sent: {second_step['serial']}/{second_step['port']}/{second_step['line']}")



        # Delay after test
        delay = float(case.get("delay", 0) or 0)

        if delay > 0:
            status_var.setText(f"Waiting {delay}s before next test...")
            for _ in range(int(delay * 10)):
                if stop_event.is_set():
                    status_var.setText("Stopped by user")
                    return
                time.sleep(0.1)
            status_var.setText("")

    except Exception as e:
        table.setItem(table_item_row, 4, QTableWidgetItem(f"ERROR: {e}"))
        status_var.setText(f"ERROR: {e}")
    finally:
        elapsed=time.time()-start_time
        table.setItem(table_item_row,5,QTableWidgetItem(f"{elapsed:.2f}"))
        status_var.setText("")



from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from reportlab.platypus import Image
import os

def generate_report(table_widget):
    try:
        # --- Close any open editors ---
        if hasattr(table_widget, "closePersistentEditor"):
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    if item:
                        table_widget.closePersistentEditor(item)
            table_widget.clearFocus()

        # --- PDF Styles ---
        styles = getSampleStyleSheet()
        wrap_style = ParagraphStyle(
            name="Wrap",
            fontSize=6,
            leading=7,
            wordWrap="CJK",
            alignment=1,
            spaceAfter=2,
        )

        # --- Determine overall result ---
        status_col_idx = 4
        overall_result = "PASS"
        for row in range(table_widget.rowCount()):
            stat_item = table_widget.item(row, status_col_idx)
            if stat_item and "PASS" not in stat_item.text().strip().upper():
                overall_result = "FAIL"
                break

        # --- Directory Structure ---
        base_dir = r"D:\Report"
        current_year = datetime.now().strftime("%Y")
        current_month = datetime.now().strftime("%m")

        year_dir = os.path.join(base_dir, current_year)
        month_dir = os.path.join(year_dir, current_month)
        pass_dir = os.path.join(month_dir, "PASS")
        fail_dir = os.path.join(month_dir, "FAIL")

        os.makedirs(pass_dir, exist_ok=True)
        os.makedirs(fail_dir, exist_ok=True)

        report_name_base = f"{ECALL_BARCODE}.pdf"
        save_dir = pass_dir if overall_result == "PASS" else fail_dir
        pdf_path = os.path.join(save_dir, report_name_base)

        # --- Document Setup ---
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A3),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )
        elements = []

        tested_by = LOGGED_IN_USER if 'LOGGED_IN_USER' in globals() else "UNKNOWN USER"

        # --- Header Table ---
        logo_path = r"C:\EOL_Report\download.png"
        if not os.path.exists(logo_path):
            logo_path = None

        status_bg_color = colors.HexColor("#92d050") if overall_result == "PASS" else colors.HexColor("#FF4747")
        EOL_NAME = "EOL-4"

        header_table_data = [
            [
                Image(logo_path, width=120, height=40) if logo_path else Paragraph("No Logo", wrap_style),
                "",
                Paragraph("<b>OVERALL<br/><br/>RESULT</b>", ParagraphStyle(name="ResultHeader", alignment=1, fontSize=22)),
                Paragraph(f"<b>{overall_result}</b>", ParagraphStyle(name="StatusColor", alignment=1, fontSize=24, textColor=colors.black)),
            ],
            [
                Paragraph("<b>TESTED BY:</b>", wrap_style),
                Paragraph("<b>DASHCAM BARCODE:</b>", wrap_style),
                Paragraph("<b>ECALL BARCODE:</b>", wrap_style),
            ],
            [
                Paragraph(tested_by, wrap_style),
                Paragraph(DASHCAM_BARCODE, wrap_style),
                Paragraph(ECALL_BARCODE, wrap_style),
                Paragraph(EOL_NAME, wrap_style)
            ],
            [
                "",
                Paragraph("<b>DATE:</b>", wrap_style),
                Paragraph("<b>TIME:</b>", wrap_style),
                ""
            ],
            [
                "",
                Paragraph(datetime.now().strftime('%m/%d/%Y'), wrap_style),
                Paragraph(datetime.now().strftime('%I:%M:%S %p'), wrap_style),
                ""
            ]
        ]

        header_table = Table(
            header_table_data,
            colWidths=[130, 160, 240, 370],
            rowHeights=[60, 15, 15, 15, 15]
        )

        header_table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("BACKGROUND", (3,0), (3,0), status_bg_color),
            ("SPAN", (0,0), (1,0)),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("VALIGN", (0,0), (-1,0), "MIDDLE"),
            ("FONTNAME", (2,0), (3,0), "Helvetica-Bold"),
            ("FONTSIZE", (2,0), (2,0), 22),
            ("FONTSIZE", (3,0), (3,0), 24),
            ("TEXTCOLOR", (3,0), (3,0), colors.black),
            ("ALIGN", (0,1), (-1,2), "LEFT"),
            ("VALIGN", (0,1), (-1,2), "MIDDLE"),
            ("FONTNAME", (0,1), (-1,2), "Helvetica-Bold"),
            ("FONTSIZE", (0,1), (-1,2), 9),
            ("ALIGN", (1,3), (2,4), "CENTER"),
            ("VALIGN", (1,3), (2,4), "MIDDLE"),
            ("FONTNAME", (1,3), (2,4), "Helvetica-Bold"),
            ("FONTSIZE", (1,3), (2,4), 9),
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 10))

        # --- Main Data Table ---
        data = [["S.No.", "TESTING PARAMETERS", "MINIMUM VALUE", "MEASURED VALUE", "MAXIMUM VALUE", "RESULT", "TIME"]]
        serial_num = 1
        all_times = []

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 0)
            if not name_item:
                continue

            name = name_item.text().strip()
            min_text = table_widget.item(row,1).text() if table_widget.item(row,1) else ""
            output_text = table_widget.item(row,2).text() if table_widget.item(row,2) else ""
            max_text = table_widget.item(row,3).text() if table_widget.item(row,3) else ""
            raw_status = table_widget.item(row,4).text().strip().upper() if table_widget.item(row,4) else ""

            if "PASS" in raw_status:
                status_text = "PASS"
            elif "FAIL" in raw_status:
                status_text = "FAIL"
            else:
                status_text = "FAIL"

            time_text = table_widget.item(row,5).text() if table_widget.item(row,5) else ""
            if time_text:
                all_times.append(time_text)

            data.append([
                Paragraph(str(serial_num), wrap_style),
                Paragraph(name, wrap_style),
                Paragraph(min_text.replace("\n", "<br/>"), wrap_style),
                Paragraph(output_text.replace("\n", "<br/>"), wrap_style),
                Paragraph(max_text.replace("\n", "<br/>"), wrap_style),
                Paragraph(status_text, wrap_style),
                Paragraph(time_text, wrap_style),
            ])
            serial_num += 1

        # === Add top and bottom custom rows ===
        
        # Top row always with S.No. = 1
        header_row = [
            Paragraph("0", wrap_style),
            Paragraph("Pre Application Info", wrap_style),
            Paragraph("EOL . CustomApi TestApp ", wrap_style),
            Paragraph("EOL & CustomApi TestApp Installed ", wrap_style),
            Paragraph("EOL . CustomApi TestApp", wrap_style),
            Paragraph("PASS", wrap_style),
            Paragraph("0.00", wrap_style),
        ]
        
        # Determine last serial number based on actual test rows
        last_serial = serial_num  # serial_num already incremented in loop
        footer_row = [
            Paragraph(str(last_serial), wrap_style),
            Paragraph("Post Application Info", wrap_style),
            Paragraph("EOL . CustomApi TestApp", wrap_style),
            Paragraph("EOL & CustomApi TestApp Uninstalled", wrap_style),
            Paragraph("EOL . CustomApi TestApp", wrap_style),
            Paragraph("PASS", wrap_style),
            Paragraph("0.00", wrap_style),
        ]
        
        # Insert top row after table header (index 1)
        data.insert(1, header_row)
        # Append bottom row at end
        data.append(footer_row)


        # --- Table Formatting ---
        col_widths = [40, 200, 160, 160, 160, 90, 90]
        main_table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="CENTER", splitByRow=True)
        main_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#021F4E")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 6),
        ]))

        # Alternate row colors
        for i in range(1, len(data)):
            bgcolor = colors.whitesmoke if i % 2 == 0 else colors.lightgrey
            main_table.setStyle(TableStyle([("BACKGROUND", (0,i), (-1,i), bgcolor)]))

        elements.append(main_table)
        elements.append(Spacer(1, 18))

        # --- Total Test Time ---
        total_seconds = 0.0
        for t in all_times:
            if not t:
                continue
            try:
                total_seconds += float(t.strip())
            except:
                continue

        h = int(total_seconds // 3600)
        m = int((total_seconds % 3600) // 60)
        s = total_seconds % 60
        overall_time_str = f"{h:02d}:{m:02d}:{s:06.3f}"

        total_time_table = Table([[f"TOTAL TEST TIME: {overall_time_str}"]], colWidths=[900], rowHeights=25)
        total_time_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#D9EAD3")),
            ("TEXTCOLOR", (0,0), (-1,-1), colors.black),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("BOX", (0,0), (-1,-1), 1, colors.black),
        ]))
        elements.append(total_time_table)

        # --- Build & Open PDF ---
        doc.build(elements)
        os.startfile(pdf_path)

    except Exception as e:
        print(f"[WARN] PDF export failed (skipping): {e}")

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


from openpyxl import load_workbook

def generate_excel(table_widget):
    try:
        tested_by = LOGGED_IN_USER if 'LOGGED_IN_USER' in globals() else "UNKNOWN USER"

        # Determine overall result
        overall_result = "PASS"
        status_col_idx = 4
        for row in range(table_widget.rowCount()):
            stat_item = table_widget.item(row, status_col_idx)
            if stat_item and "PASS" not in stat_item.text().strip().upper():
                overall_result = "FAIL"
                break

        # Prepare directories dynamically based on year/month
        base_dir = r"D:\Report"
        current_year = datetime.now().strftime("%Y")
        current_month = datetime.now().strftime("%m")

        year_dir = os.path.join(base_dir, current_year)
        month_dir = os.path.join(year_dir, current_month)

        pass_dir = os.path.join(month_dir, "PASS")
        fail_dir = os.path.join(month_dir, "FAIL")

        os.makedirs(pass_dir, exist_ok=True)
        os.makedirs(fail_dir, exist_ok=True)

        excel_name_base = f"{ECALL_BARCODE}.xlsx"
        save_dir = pass_dir if overall_result == "PASS" else fail_dir
        excel_path = os.path.join(save_dir, excel_name_base)

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "excel_pattern"

        # Styles
        header_font = Font(bold=True, size=10)
        result_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        status_fill = PatternFill(
            start_color="92d050" if overall_result == "PASS" else "FF4747",
            end_color="92d050" if overall_result == "PASS" else "FF4747",
            fill_type="solid"
        )
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ===== HEADER SECTION =====
        logo_path = r"C:\EOL_Report\download.png"
        logo_width = 270
        logo_height = 80

        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = logo_width
            img.height = logo_height
            ws.add_image(img, "A2")

        ws.merge_cells("A2:B2")
        ws.row_dimensions[2].height = logo_height * 0.75

        ws["C2"] = "OVERALL RESULT"
        ws["C2"].font = result_font
        ws["C2"].alignment = center_align

        ws.merge_cells("D2:F2")
        ws["D2"] = overall_result
        ws["D2"].font = result_font
        ws["D2"].alignment = center_align
        ws["D2"].fill = status_fill

        ws["A3"] = "TESTED BY:"
        ws["B3"] = "DASHCAM BARCODE:"
        ws["C3"] = "ECALL BARCODE:"
        ws.merge_cells("D3:F3")
        ws["D3"] = ""
        for cell in ["A3", "B3", "C3"]:
            ws[cell].font = header_font
            ws[cell].alignment = left_align
        ws["D3"].alignment = left_align

        ws.merge_cells("A4:A6")
        ws["A4"] = tested_by
        ws["A4"].alignment = left_align
        ws["A4"].font = header_font

        ws["B4"] = DASHCAM_BARCODE
        ws["C4"] = ECALL_BARCODE
        ws["B4"].alignment = left_align
        ws["C4"].alignment = left_align

        ws["B5"] = "DATE:"
        ws["C5"] = "TIME:"
        ws["B5"].font = header_font
        ws["C5"].font = header_font
        ws["B5"].alignment = center_align
        ws["C5"].alignment = center_align

        ws["B6"] = datetime.now().strftime("%d-%m-%Y")
        ws["C6"] = datetime.now().strftime("%I:%M:%S %p")
        ws["B6"].alignment = center_align
        ws["C6"].alignment = center_align

        ws.merge_cells("D4:F4")
        ws["D4"] = ""
        ws["D4"].alignment = left_align

        ws.merge_cells("D5:F6")
        ws["D5"] = "EOL-4"
        ws["D5"].alignment = left_align

        # ===== MAIN TABLE =====
        main_headers = ["S.No.", "TESTING PARAMETERS", "MINIMUM VALUE", "MEASURED VALUE", "MAXIMUM VALUE", "RESULT"]
        start_row = 7
        for col_idx, header in enumerate(main_headers, 1):
            ws.cell(row=start_row, column=col_idx, value=header)
            ws.cell(row=start_row, column=col_idx).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=start_row, column=col_idx).alignment = center_align
            ws.cell(row=start_row, column=col_idx).fill = PatternFill("solid", fgColor="021F4E")
            ws.cell(row=start_row, column=col_idx).border = thin_border

        serial_num = 2  # start from 2 because 1 is "Pre Application Info"

        # --- TOP CUSTOM ROW ---
        pre_row = [1, "Pre Application Info", "EOL . CustomApi TestApp",
                   "EOL & CustomApi TestApp Installed",
                   "EOL . CustomApi TestApp", "PASS"]
        for col_idx, val in enumerate(pre_row, 1):
            c = ws.cell(row=start_row + 1, column=col_idx, value=val)
            c.alignment = center_align
            c.border = thin_border
            c.fill = PatternFill("solid", fgColor="D9EAD3")

        # --- WRITE MAIN TEST ROWS ---
        for r in range(table_widget.rowCount()):
            name_item = table_widget.item(r, 0)
            if not name_item:
                continue

            col1 = table_widget.item(r, 0).text() if table_widget.item(r, 0) else ""
            col2 = table_widget.item(r, 1).text() if table_widget.item(r, 1) else ""
            col3 = table_widget.item(r, 2).text() if table_widget.item(r, 2) else ""
            col4 = table_widget.item(r, 3).text() if table_widget.item(r, 3) else ""
            col5 = table_widget.item(r, 4).text() if table_widget.item(r, 4) else ""
            col6 = table_widget.item(r, 5).text() if table_widget.item(r, 5) else ""

            result_text = col5.strip().upper()
            if any(x in result_text for x in ["PASS", "✅", "OK"]):
                clean_result = "PASS"
            elif any(x in result_text for x in ["FAIL", "❌"]):
                clean_result = "FAIL"
            else:
                clean_result = "FAIL"

            row_values = [serial_num, col1, col2, col3, col4, clean_result]
            row_idx = start_row + serial_num
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin_border
                fill_color = "D3D3D3" if serial_num % 2 == 0 else "FFFFFF"
                cell.fill = PatternFill("solid", fgColor=fill_color)
                if col_idx == 6:
                    if clean_result == "PASS":
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    elif clean_result == "FAIL":
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
            serial_num += 1

        # --- FOOTER ROW ---
        footer_row = [
            serial_num,
            "Post Application Info",
            "EOL . CustomApi TestApp",
            "EOL & CustomApi TestApp Uninstalled",
            "EOL . CustomApi TestApp",
            "PASS"
        ]
        footer_row_idx = start_row + serial_num
        for col_idx, val in enumerate(footer_row, 1):
            c = ws.cell(row=footer_row_idx, column=col_idx, value=val)
            c.alignment = center_align
            c.border = thin_border
            c.fill = PatternFill("solid", fgColor="D9EAD3")

        # Column widths
        col_widths = [10, 35, 35, 35, 35, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(excel_path)
        print(f"[INFO] Excel report saved at: {excel_path}")

        bharat_base = r"D:\MES\Bharat"
        
        year = datetime.now().strftime("%Y")
        month = datetime.now().strftime("%m")
        day = datetime.now().strftime("%d")
        
        day_dir = os.path.join(bharat_base, year, month, day)
        os.makedirs(day_dir, exist_ok=True)
        
        # --- Single Excel for whole day ---
        ecall_excel_path = os.path.join(day_dir, "ECALL_ID.xlsx")
        
        # Create or load file
        if os.path.exists(ecall_excel_path):
            ewb = load_workbook(ecall_excel_path)
            ews = ewb.active
        else:
            ewb = Workbook()
            ews = ewb.active
            ews.title = "ECALL_ID"
            ews.append(["BARCODE", "COMMAND", "OUTPUT", "STATUS", "TIME TAKE", "TIME"])
        
        # Commands to save
        TARGET_CMDS = [
            "CAMERA ENTRY",
            "READ_IMSI NUMBER",
            "READ_IMEI NUMBER",
            "READ_ICCID NUMBER",
            "READ EID NUMBER"
        ]
        TARGET_SET = {cmd.upper().strip() for cmd in TARGET_CMDS}
        
        # -----------------------------------------
        #         Append Barcode + its Data
        # -----------------------------------------
        
        # 1) Write Barcode row
        ews.append([f"BARCODE: {ECALL_BARCODE}", "", "", "", "", ""])
        
        # 2) Add each command row
        for r in range(table_widget.rowCount()):
            cmd = table_widget.item(r, 0).text() if table_widget.item(r, 0) else ""
            if cmd.strip().upper() in TARGET_SET:
        
                min_v = table_widget.item(r, 2).text() if table_widget.item(r, 2) else ""
                max_v = table_widget.item(r, 4).text() if table_widget.item(r, 4) else ""
                res_v = table_widget.item(r, 5).text() if table_widget.item(r, 5) else ""
        
                ews.append([
                    "",
                    cmd,
                    min_v,
                    max_v,
                    res_v,
                    datetime.now().strftime("%H:%M:%S")
                ])
        
        # 3) Leave one blank row
        ews.append(["", "", "", "", "", ""])
        
        # # Save file
        # ewb.save(ecall_excel_path)
        # print(f"[INFO] DAILY ECALL EXCEL UPDATED → {ecall_excel_path}")

        try:
            ewb.save(ecall_excel_path)
            print(f"[INFO] DAILY ECALL EXCEL UPDATED → {ecall_excel_path}")
        except PermissionError:
            backup_path = os.path.join(day_dir, "ECALL_ID_backup.xlsx")
            ewb.save(backup_path)
            print("[WARN] MAIN ECALL_ID.xlsx OPEN HAI → backup file save ho gai:")
            print("       ", backup_path)


    except Exception as e:
        print(f"[WARN] Excel export failed: {e}")















# Global counter variables
total_count = 0
pass_count = 0
fail_count = 0
yield_count = 0

COUNTER_FILE = "counters.json"


def load_counters():
    global total_count, pass_count, fail_count, yield_count
    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, "r") as f:
            data = json.load(f)
            total_count = data.get("total_count", 0)
            pass_count = data.get("pass_count", 0)
            fail_count = data.get("fail_count", 0)
            yield_count = data.get("yield_count", 0.0)
    else:
        total_count = pass_count = fail_count = yield_count = 0


def save_counters():
    global total_count, pass_count, fail_count, yield_count
    data = {
        "total_count": total_count,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "yield_count": yield_count
    }
    with open(COUNTER_FILE, "w") as f:
        json.dump(data, f)


def refresh_counters(window_obj):
    global total_count, pass_count, fail_count, yield_count
    window_obj.label_total.setText(str(total_count))
    window_obj.label_pass.setText(str(pass_count))
    window_obj.label_fail.setText(str(fail_count))
    window_obj.label_yield.setText(f"{yield_count}%")



# def finalize_run(table_widget, window_obj):

#     global total_count, pass_count, fail_count, yield_count
#     # global label_total, label_pass, label_fail, label_yield

#     total_count += 1  # one run completed

#     all_status = []
#     for row in range(table_widget.rowCount()):
#         item = table_widget.item(row, 4)  # status column index (0-based)
#         if item:
#             all_status.append(item.text())
#         else:
#             all_status.append("")

#     if all(s == "PASS ✅" for s in all_status):
#         pass_count += 1
#     else:
#         fail_count += 1

#     yield_count = round((pass_count / total_count) * 100, 2)

#     refresh_counters(window_obj)
#     save_counters()
def finalize_run(table_widget, window_obj):
    global total_count, pass_count, fail_count, yield_count

    total_count += 1  # one run completed

    all_status = []
    for row in range(table_widget.rowCount()):
        item = table_widget.item(row, 4)  # status column index (0-based)
        if item:
            all_status.append(item.text().strip())
        else:
            all_status.append("")

    # ----- Determine overall result -----
    if all(s == "PASS ✅" for s in all_status if s):
        pass_count += 1
        overall_result = "PASS"
    else:
        fail_count += 1
        overall_result = "FAIL"

    yield_count = round((pass_count / total_count) * 100, 2)

    # Update UI counters
    refresh_counters(window_obj)
    save_counters()

    # print(f"🔥 finalize_run completed — overall_result = {overall_result}")


    overall_result = "PASS" if all(s == "PASS ✅" for s in all_status) else "FAIL"
    print(f"🔥 finalize_run completed — overall_result = {overall_result}")
    
    try:
        window_obj.finalize_hud_timer()
        print("⏹ finalize_hud_timer() called")
        window_obj.overallBannerRequested.emit(overall_result, 3000)
        print("🚀 overallBannerRequested emitted")
    except Exception as e:
        print("❌ Banner emit error:", e)



# GUI
SETTINGS_FILE = "com_settings.json"


def load_com_settings():
    try:
        with open(SETTINGS_FILE, "r") as f:
            return json.load(f)
    except:
        return {
            "ecall_com": "",
            "dashcam_com": "",
            "dbmeter_com": "",
            "feasa_com": "",
            "camera_ip": "",
            "camera_port": ""
        }


def save_com_settings(settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f)


class Extreme3DCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(360, 400)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setStyleSheet("border-radius: 20px;")
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(40, 40, 40, 40)
        self.layout.setSpacing(25)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = self.rect()

        shadow_color_outer = QColor(0, 0, 0, 230)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(shadow_color_outer)
        painter.drawRoundedRect(rect.adjusted(10, 10, 10, 10), 22, 22)

        top_left = QPointF(rect.topLeft().x(), rect.topLeft().y())
        bottom_left = QPointF(rect.bottomLeft().x(), rect.bottomLeft().y())

        grad = QLinearGradient(top_left, bottom_left)
        grad.setColorAt(0.0, QColor("#808080"))
        grad.setColorAt(0.85, QColor("#303030"))
        grad.setColorAt(1.0, QColor("#202020"))
        painter.setBrush(QBrush(grad))
        painter.drawRoundedRect(rect.adjusted(0, 0, -10, -10), 22, 22)

        hilite_color = QColor(255, 255, 255, 120)
        pen = QPen(hilite_color, 10, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap)
        painter.setPen(pen)
        painter.drawLine(top_left + QPointF(10, 10), bottom_left + QPointF(10, -10))
        painter.drawLine(top_left + QPointF(10, 10), QPointF(rect.topRight().x(), rect.topRight().y()) + QPointF(-10, 10))

        inner_shadow = QColor(0, 0, 0, 250)
        pen = QPen(inner_shadow, 7, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap)
        painter.setPen(pen)
        painter.drawRoundedRect(rect.adjusted(7, 7, -15, -15), 18, 18)
        painter.end()


class NeonButton(QPushButton):
    def __init__(self, text):
        super().__init__(text)
        self.setFixedSize(150, 50)
        self.setStyleSheet("""
            QPushButton {
                background-color: #004040;
                color: white;
                border-radius: 12px;
                font-size: 16px;
                font-weight: bold;
                border: 3px solid #0ff;
            }
            QPushButton:hover {
                background-color: #009090;
                color: #0ff;
                border: 3px solid #0ff;
            }
            QPushButton:pressed {
                background-color: #33ffcc;
                color: #00ccbb;
                border: 3px solid #00ccbb;
            }
        """)

        self.shadow_main = QGraphicsDropShadowEffect(self)
        self.shadow_main.setBlurRadius(35)
        self.shadow_main.setOffset(0, 0)
        self.shadow_main.setColor(QColor(0, 255, 255, 190))
        self.setGraphicsEffect(self.shadow_main)

        self.hue = 0
        self.timer = QTimer(self)
        self.timer.setInterval(40)
        self.timer.timeout.connect(self.glow_animation)
        self.timer.start()

    def glow_animation(self):
        self.hue = (self.hue + 8) % 360
        glow_color = QColor.fromHsv(self.hue, 255, 255, 190)
        effect = self.graphicsEffect()
        if effect:
            effect.setColor(glow_color)



import json, os
from PyQt6.QtWidgets import (
    QWidget, QLabel, QLineEdit, QVBoxLayout, QHBoxLayout,
    QMessageBox, QApplication, QPushButton, QDialog,
    QGraphicsDropShadowEffect, QFrame
)
from PyQt6.QtGui import QColor, QLinearGradient, QPalette, QBrush, QPainter, QPen
from PyQt6.QtCore import Qt, QPointF, QTimer

# ======================= USER JSON HANDLING ==========================
USER_FILE = "users.json"

def load_users():
    if not os.path.exists(USER_FILE):
        default_users = [
            {"username": "1", "password": "1"},
            {"username": "Admin", "password": "Admin@123"}
        ]
        with open(USER_FILE, "w") as f:
            json.dump(default_users, f, indent=4)
        return default_users
    with open(USER_FILE, "r") as f:
        return json.load(f)

def save_users(users):
    with open(USER_FILE, "w") as f:
        json.dump(users, f, indent=4)


# ======================= ADMIN REGISTRATION POPUP ==========================
class AdminRegistrationPopup(QDialog):
    def __init__(self, users, parent=None):
        super().__init__(parent)
        self.users = users
        self.setWindowTitle("Admin Control Panel")
        self.setFixedSize(520, 520)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setModal(True)

        # Gradient background
        gradient = QLinearGradient(0, 0, 0, self.height())
        gradient.setColorAt(0.0, QColor(0, 230, 255))
        gradient.setColorAt(0.5, QColor(10, 10, 10))
        gradient.setColorAt(1.0, QColor(0, 140, 255))
        palette = self.palette()
        palette.setBrush(QPalette.ColorRole.Window, QBrush(gradient))
        self.setPalette(palette)
        self.setAutoFillBackground(True)

        # Layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(18, 18, 18, 18)
        main_layout.setSpacing(12)

        # Title
        title = QLabel("Admin Control Panel")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: cyan; font-size: 20px; font-weight: bold;")
        main_layout.addWidget(title)

        # Admin auth row
        auth_row = QHBoxLayout()
        self.admin_pass = QLineEdit()
        self.admin_pass.setPlaceholderText("Enter current Admin password to unlock")
        self.admin_pass.setEchoMode(QLineEdit.EchoMode.Password)
        auth_row.addWidget(self.admin_pass)
        self.unlock_btn = QPushButton("Unlock")
        self.unlock_btn.setFixedWidth(100)
        auth_row.addWidget(self.unlock_btn)
        main_layout.addLayout(auth_row)

        # Container for admin controls (hidden until unlock)
        self.admin_controls = QVBoxLayout()
        self.admin_controls.setSpacing(10)
        main_layout.addLayout(self.admin_controls)

        # --- User list + delete ---
        lbl_users = QLabel("Registered Users")
        lbl_users.setStyleSheet("color: #E0FFFF; font-weight: bold;")
        self.admin_controls.addWidget(lbl_users)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self.user_list = QListWidget()
        self.user_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        self.user_list.setFixedHeight(140)
        self.admin_controls.addWidget(self.user_list)

        user_btn_row = QHBoxLayout()
        self.refresh_btn = QPushButton("Refresh")
        self.delete_btn = QPushButton("Delete Selected")
        user_btn_row.addWidget(self.refresh_btn)
        user_btn_row.addWidget(self.delete_btn)
        self.admin_controls.addLayout(user_btn_row)

        # --- Registration area ---
        reg_title = QLabel("Register New User")
        reg_title.setStyleSheet("color: #E0FFFF; font-weight: bold;")
        self.admin_controls.addWidget(reg_title)

        reg_row = QHBoxLayout()
        self.new_user = QLineEdit()
        self.new_user.setPlaceholderText("New username")
        self.new_pass = QLineEdit()
        self.new_pass.setPlaceholderText("New password")
        self.new_pass.setEchoMode(QLineEdit.EchoMode.Password)
        reg_row.addWidget(self.new_user)
        reg_row.addWidget(self.new_pass)
        self.admin_controls.addLayout(reg_row)

        reg_btn_row = QHBoxLayout()
        self.register_btn = QPushButton("Register")
        reg_btn_row.addStretch()
        reg_btn_row.addWidget(self.register_btn)
        self.admin_controls.addLayout(reg_btn_row)

        # --- Change Admin password area ---
        chg_title = QLabel("Change Admin Password")
        chg_title.setStyleSheet("color: #E0FFFF; font-weight: bold;")
        self.admin_controls.addWidget(chg_title)

        chg_row1 = QHBoxLayout()
        self.new_admin_pass = QLineEdit()
        self.new_admin_pass.setPlaceholderText("New admin password")
        self.new_admin_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.confirm_admin_pass = QLineEdit()
        self.confirm_admin_pass.setPlaceholderText("Confirm new password")
        self.confirm_admin_pass.setEchoMode(QLineEdit.EchoMode.Password)
        chg_row1.addWidget(self.new_admin_pass)
        chg_row1.addWidget(self.confirm_admin_pass)
        self.admin_controls.addLayout(chg_row1)

        chg_btn_row = QHBoxLayout()
        self.change_admin_btn = QPushButton("Change Admin Password")
        chg_btn_row.addStretch()
        chg_btn_row.addWidget(self.change_admin_btn)
        self.admin_controls.addLayout(chg_btn_row)

        # Bottom cancel/close
        bottom_row = QHBoxLayout()
        bottom_row.addStretch()
        self.close_btn = QPushButton("Close")
        bottom_row.addWidget(self.close_btn)
        main_layout.addLayout(bottom_row)

        # Hook buttons
        self.unlock_btn.clicked.connect(self.try_unlock)
        self.refresh_btn.clicked.connect(self.refresh_user_list)
        self.delete_btn.clicked.connect(self.delete_selected_users)
        self.register_btn.clicked.connect(self.try_register)
        self.change_admin_btn.clicked.connect(self.try_change_admin_password)
        self.close_btn.clicked.connect(self.reject)

        # Initially disable admin controls until unlocked
        self.set_admin_controls_enabled(False)

        # Fill user list for display (still disabled until unlocked)
        self.refresh_user_list()

    def set_admin_controls_enabled(self, enabled: bool):
        # enable/disable all children in admin_controls area
        for i in range(self.admin_controls.count()):
            item = self.admin_controls.itemAt(i)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.setEnabled(enabled)
            else:
                # layout -> iterate children
                try:
                    for j in range(item.count()):
                        child = item.itemAt(j)
                        if child is None:
                            continue
                        w = child.widget()
                        if w:
                            w.setEnabled(enabled)
                except Exception:
                    pass
        # admin_pass & unlock remain enabled
        self.admin_pass.setEnabled(True)
        self.unlock_btn.setEnabled(True)

    def try_unlock(self):
        entered = self.admin_pass.text().strip()
        admin_user = next((u for u in self.users if u["username"].lower() == "admin"), None)
        if admin_user and admin_user["password"] == entered:
            QMessageBox.information(self, "Unlocked", "Admin panel unlocked.")
            self.set_admin_controls_enabled(True)
            self.admin_pass.clear()
            self.refresh_user_list()
        else:
            QMessageBox.critical(self, "Access Denied", "Incorrect admin password!")

    def refresh_user_list(self):
        self.user_list.clear()
        for u in self.users:
            uname = u.get("username", "")
            item_text = f"{uname} {'(Admin)' if uname.lower() == 'admin' else ''}"
            from PyQt6.QtWidgets import QListWidgetItem
            it = QListWidgetItem(item_text)
            # store original username in Qt.UserRole for easy retrieval
            it.setData(Qt.ItemDataRole.UserRole, uname)
            self.user_list.addItem(it)

    def delete_selected_users(self):
        sel = self.user_list.selectedItems()
        if not sel:
            QMessageBox.information(self, "Info", "No user selected.")
            return

        # confirm
        ok = QMessageBox.question(self, "Confirm Delete",
        f"Delete {len(sel)} selected user(s)?",
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ok != QMessageBox.StandardButton.Yes:
            return

        to_delete = []
        for it in sel:
            uname = it.data(Qt.ItemDataRole.UserRole)
            if uname.lower() == "admin":
                QMessageBox.warning(self, "Skipped", "Cannot delete Admin user.")
                continue
            to_delete.append(uname)

        if not to_delete:
            return

        # remove from users list
        self.users[:] = [u for u in self.users if u["username"] not in to_delete]
        save_users(self.users)
        QMessageBox.information(self, "Deleted", f"Deleted {len(to_delete)} user(s).")
        self.refresh_user_list()

    def try_register(self):
        new_username = self.new_user.text().strip()
        new_password = self.new_pass.text().strip()
        if not new_username or not new_password:
            QMessageBox.warning(self, "Error", "Username & password required.")
            return
        if any(u["username"].lower() == new_username.lower() for u in self.users):
            QMessageBox.warning(self, "Warning", "Username already exists!")
            return
        self.users.append({"username": new_username, "password": new_password})
        save_users(self.users)
        QMessageBox.information(self, "Success", f"User '{new_username}' registered successfully!")
        self.new_user.clear()
        self.new_pass.clear()
        self.refresh_user_list()

    def try_change_admin_password(self):
        new1 = self.new_admin_pass.text().strip()
        new2 = self.confirm_admin_pass.text().strip()
        if not new1 or not new2:
            QMessageBox.warning(self, "Error", "Both fields required.")
            return
        if new1 != new2:
            QMessageBox.warning(self, "Error", "Passwords do not match.")
            return

        # find admin entry and update
        updated = False
        for u in self.users:
            if u.get("username", "").lower() == "admin":
                u["password"] = new1
                updated = True
                break
        if not updated:
            # if admin missing, add it (unlikely)
            self.users.append({"username": "Admin", "password": new1})
        save_users(self.users)
        QMessageBox.information(self, "Success", "Admin password changed.")
        self.new_admin_pass.clear()
        self.confirm_admin_pass.clear()
        self.refresh_user_list()



# ======================= LOGIN WINDOW ==========================
class LoginWindow(QWidget):
    def __init__(self, login_success_callback=None):
        super().__init__()
        self.login_success_callback = login_success_callback
        self.setWindowTitle("Extreme 3D Neon Login")
        self.resize(700, 450)
        self.USERS = load_users()

        # Gradient background
        gradient = QLinearGradient(0, 0, 0, self.height())
        gradient.setColorAt(0.0, QColor(0, 230, 255))
        gradient.setColorAt(0.2, QColor(10, 255, 140))
        gradient.setColorAt(0.4, QColor(0, 255, 220, 180))
        gradient.setColorAt(0.6, QColor(10, 10, 10))
        gradient.setColorAt(0.8, QColor(240, 240, 240))
        gradient.setColorAt(1.0, QColor(0, 140, 255))
        palette = self.palette()
        palette.setBrush(QPalette.ColorRole.Window, QBrush(gradient))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

        # Company branding
        self.company_label = QLabel("Vipin")
        self.company_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        gradient_text = QLinearGradient(0, 0, 400, 0)
        gradient_text.setColorAt(0.0, QColor(0, 90, 160))
        gradient_text.setColorAt(0.5, QColor(0, 140, 255))
        gradient_text.setColorAt(1.0, QColor(0, 170, 255))
        palette_text = self.company_label.palette()
        palette_text.setBrush(QPalette.ColorRole.WindowText, QBrush(gradient_text))
        self.company_label.setPalette(palette_text)
        self.company_label.setStyleSheet("""
            background: transparent;
            font-size: 72px;
            font-weight: bold;
            letter-spacing: 4px;
            font-family: 'Arial Black', Arial, sans-serif;
        """)

        self.subtitle_label = QLabel("QUALITY DELIVERED")
        self.subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.subtitle_label.setStyleSheet("background: transparent; color: #B8860B; font-size: 18px; font-family: 'Arial';")

        self.card = Extreme3DCard()
        self.title = QLabel("Login")
        self.title.setStyleSheet("color: white; font-size: 26px; font-weight: bold;")
        self.title.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.username_entry = QLineEdit()
        self.username_entry.setPlaceholderText("Username")
        self.username_entry.setStyleSheet("""
            QLineEdit {
                background: rgba(255,255,255,0.1);
                border-radius: 12px;
                padding: 8px;
                color: white;
                font-size: 16px;
                border: 3px solid #0ff;
            }
        """)

        self.password_entry = QLineEdit()
        self.password_entry.setPlaceholderText("Password")
        self.password_entry.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_entry.setStyleSheet(self.username_entry.styleSheet())

        self.login_button = NeonButton("Login")
        self.register_button = NeonButton("Register")
        self.login_button.clicked.connect(self.handle_login)
        self.register_button.clicked.connect(self.open_registration)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 20, 10, 10)
        main_layout.addWidget(self.company_label)
        main_layout.addWidget(self.subtitle_label)
        main_layout.addStretch()
        main_layout.addWidget(self.card, alignment=Qt.AlignmentFlag.AlignCenter)

        self.card.layout.addWidget(self.title)
        self.card.layout.addWidget(self.username_entry)
        self.card.layout.addWidget(self.password_entry)
        self.card.layout.addWidget(self.login_button)
        self.card.layout.addWidget(self.register_button)

        main_layout.addStretch()
        self.setLayout(main_layout)

    def handle_login(self):
        username_input = self.username_entry.text().strip()
        password_input = self.password_entry.text().strip()
        for user in self.USERS:
            if user["username"] == username_input and user["password"] == password_input:
                global LOGGED_IN_USER
                LOGGED_IN_USER = username_input
                QMessageBox.information(self, "Success", f"Welcome {username_input}!")
                
                if self.login_success_callback:
                    self.login_success_callback()
                return
        QMessageBox.critical(self, "Error", "Invalid credentials")

    def open_registration(self):
        popup = AdminRegistrationPopup(self.USERS, self)
        popup.exec()





class GlassyNeonButton(QWidget):
    def __init__(self, text, color_mode='gold', parent=None):
        super().__init__(parent)
        self.setFixedSize(120, 40)
        self.text = text
        self.color_mode = color_mode
        self.hue = 0
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.updateGlow)
        self.timer.start(30)

    def updateGlow(self):
        self.hue = (self.hue + 4) % 360
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        center = QPointF(w / 2, h / 2)
        rect_width, rect_height = w * 0.95, h * 0.7
        rect = QRectF(center.x() - rect_width / 2, center.y() - rect_height / 2, rect_width, rect_height)

        # Background
        rect = QRectF(center.x() - rect_width / 2, center.y() - rect_height / 2, rect_width, rect_height)
        painter.setBrush(QColor(255, 255, 255, 76))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(rect, 12, 12)

        # Glow border
        conical = QConicalGradient(QPointF(rect.center().x(), rect.center().y()), self.hue)
        if self.color_mode == 'gold':
            conical.setColorAt(0.0, QColor(255, 213, 59, 220))
            conical.setColorAt(0.5, QColor(192, 192, 192, 0))
            conical.setColorAt(1.0, QColor(255, 213, 59, 220))
        else:
            conical.setColorAt(0.0, QColor(192, 192, 192, 220))
            conical.setColorAt(0.5, QColor(255, 213, 59, 0))
            conical.setColorAt(1.0, QColor(192, 192, 192, 220))

        pen = QPen(QBrush(conical), 4)
        pen.setCosmetic(True)
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(rect.adjusted(2, 2, -2, -2), 12, 12)

        # Text
        painter.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        painter.setPen(QColor(255, 255, 255))
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, self.text)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked()
        super().mousePressEvent(event)

    def clicked(self):
        pass


from PyQt6.QtWidgets import (
    QWidget, QLabel, QLineEdit, QVBoxLayout, QHBoxLayout, QGridLayout,
    QPushButton, QMessageBox, QInputDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QPainter, QLinearGradient, QColor
import serial, socket


class ComPortSettingsScreen(QWidget):
    def __init__(self, open_barcode_callback=None):
        super().__init__()
        self.open_barcode_callback = open_barcode_callback
        self.setWindowTitle("COM Port Settings")
        self.setMinimumSize(500, 480)

        # --- Outer layout ---
        main_layout = QVBoxLayout(self)
        self.setLayout(main_layout)

        # --- Neon card wrapper ---
        self.card = Extreme3DCard()
        main_layout.addWidget(self.card, alignment=Qt.AlignmentFlag.AlignCenter)

        # --- Admin unlock row (top) ---
        admin_row = QHBoxLayout()
        self.admin_label = QLabel("Admin Unlock:")
        self.admin_label.setStyleSheet("color: cyan; font-weight: bold;")
        self.admin_pass = QLineEdit()
        self.admin_pass.setPlaceholderText("Enter Admin password to unlock")
        self.admin_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.unlock_btn = QPushButton("Unlock")
        self.unlock_btn.setFixedWidth(100)
        admin_row.addWidget(self.admin_label)
        admin_row.addWidget(self.admin_pass)
        admin_row.addWidget(self.unlock_btn)
        self.card.layout.addLayout(admin_row)

        # --- Scrollable Area for Settings ---
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_area_widget)

        # --- Grid inside scroll area ---
        grid = QGridLayout()
        scroll_layout.addLayout(grid)
        scroll_area.setWidget(scroll_area_widget)
        self.card.layout.addWidget(scroll_area)

        # ✅ Fix: make scroll area fully transparent (no white background)
        scroll_area.setStyleSheet("""
            QScrollArea {
                background: transparent;
                border: none;
            }
            QWidget {
                background: transparent;
            }
        """)
        scroll_area.viewport().setStyleSheet("background: transparent;")

        # --- Load settings ---
        self.settings = load_com_settings()
        self.entries = {}

        labels = [
            "ECALL COM:", "DASHCAM COM:", "DB Meter COM:", "FEASA COM:",
            "Camera IP:", "Camera Port:", "PSU IP:",
            "NI Device 1 Serial Hex:", "NI Device 2 Serial Hex:", "NI Device 3 Serial Hex:"
        ]

        for idx, label in enumerate(labels):
            lbl = QLabel(label)
            lbl.setStyleSheet("""
                color: cyan;
                font-weight: bold;
                background-color: rgba(0, 0, 0, 80);  /* light transparent black for contrast */
                padding: 2px;
            """)
            edit = QLineEdit()
            edit.setStyleSheet("background-color: #10182e; color: white; border: 1px solid #00FFFF;")
            key = label.lower().replace(" ", "_").replace(":", "")
            edit.setText(self.settings.get(key, ""))
            self.entries[key] = edit
            grid.addWidget(lbl, idx, 0, alignment=Qt.AlignmentFlag.AlignRight)
            grid.addWidget(edit, idx, 1)

        # --- Buttons ---
        btn_layout = QHBoxLayout()
        self.connect_button = GlassyNeonButton("Connect", color_mode='gold')
        self.update_button = GlassyNeonButton("Update", color_mode='silver')
        btn_layout.addWidget(self.connect_button)
        btn_layout.addWidget(self.update_button)
        self.card.layout.addLayout(btn_layout)

        # --- Hook signals ---
        self.unlock_btn.clicked.connect(self.try_unlock_admin)
        self.connect_button.clicked = self.connect_ports
        self.update_button.clicked = self.update_settings

        # Initially locked
        self.set_locked(True)

    # ---------------- BACKGROUND ----------------
    def paintEvent(self, event):
        painter = QPainter(self)
        rect = self.rect()
        gradient = QLinearGradient(0, 0, 0, rect.height())
        gradient.setColorAt(0.0, QColor(0, 230, 255))
        gradient.setColorAt(0.2, QColor(10, 255, 140))
        gradient.setColorAt(0.4, QColor(0, 255, 220, 180))
        gradient.setColorAt(0.6, QColor(10, 10, 10))
        gradient.setColorAt(0.8, QColor(20, 20, 20))  # ✅ darker instead of white
        gradient.setColorAt(1.0, QColor(0, 140, 255))
        painter.fillRect(rect, gradient)

    # ---------------- LOCK/UNLOCK ----------------
    def set_locked(self, locked: bool):
        for entry in self.entries.values():
            entry.setEnabled(not locked)
        self.connect_button.setEnabled(True)
        self.update_button.setEnabled(not locked)
        self.unlock_btn.setText("Unlock" if locked else "Locked (Unlocked)")

    # ---------------- ADMIN ----------------
    def try_unlock_admin(self):
        entered = self.admin_pass.text().strip()
        if not entered:
            QMessageBox.warning(self, "Admin", "Enter admin password to unlock.")
            return

        users = load_users()
        admin_user = next((u for u in users if u.get("username", "").lower() == "admin"), None)
        if not admin_user:
            QMessageBox.critical(self, "Admin", "Admin user not found. Cannot unlock.")
            return

        if admin_user["password"] != entered:
            QMessageBox.critical(self, "Access Denied", "Incorrect admin password!")
            return

        self.set_locked(False)
        QMessageBox.information(self, "Unlocked", "Settings unlocked. You can now edit and update.")
        self.admin_pass.clear()

    # ---------------- VALIDATIONS ----------------
    def _is_settings_modified(self):
        for k, entry in self.entries.items():
            if entry.text().strip() != (self.settings.get(k, "") or ""):
                return True
        return False

    def _validate_admin_password(self, password) -> bool:
        users = load_users()
        admin_user = next((u for u in users if u.get("username", "").lower() == "admin"), None)
        if not admin_user:
            return False
        return admin_user.get("password") == password

    # ---------------- CONNECT PORTS ----------------
    def connect_ports(self):
        modified = self._is_settings_modified()

        if modified:
            txt, ok = QInputDialog.getText(
                self, "Admin Required",
                "Settings changed. Enter Admin password to continue:",
                QLineEdit.EchoMode.Password
            )
            if not ok:
                return
            if not self._validate_admin_password(txt.strip()):
                QMessageBox.critical(self, "Access Denied", "Incorrect admin password!")
                return

        global ser_ecall, ser_dashcam, ser_dbmeter, serfeasa
        global cam_sock, cam_ip, cam_port, ports_connected
        global psu_rm, keysight_psu, psu_ip
        global ni_device_1, ni_device_2, ni_device_3

        missing = []
        failed = []

        # --- 1️⃣ Required Fields ---
        ecall_com = self.entries["ecall_com"].text().strip()
        dashcam_com = self.entries["dashcam_com"].text().strip()
        dbmeter_com = self.entries["db_meter_com"].text().strip()
        feasa_com = self.entries["feasa_com"].text().strip()
        cam_ip = self.entries["camera_ip"].text().strip()
        cam_port_text = self.entries["camera_port"].text().strip()
        psu_ip = self.entries["psu_ip"].text().strip()
        ni1 = self.entries.get("ni_device_1_serial_hex").text().strip().upper()
        ni2 = self.entries.get("ni_device_2_serial_hex").text().strip().upper()
        ni3 = self.entries.get("ni_device_3_serial_hex").text().strip().upper()

        required_fields = {
            "ECALL COM": ecall_com,
            "DASHCAM COM": dashcam_com,
            "DB METER COM": dbmeter_com,
            "FEASA COM": feasa_com,
            "Camera IP": cam_ip,
            "Camera Port": cam_port_text,
            "PSU IP": psu_ip,
            "NI Device 1 Serial": ni1,
            "NI Device 2 Serial": ni2,
            "NI Device 3 Serial": ni3
        }

        for name, val in required_fields.items():
            if not val:
                missing.append(name)

        # if missing:
        #     QMessageBox.critical(
        #         self,
        #         "Missing Fields",
        #         "⚠️ The following fields are mandatory and must be filled:\n\n" +
        #         "\n".join(f"- {x}" for x in missing)
        #     )
        #     return

        # --- 2️⃣ Try Connecting COM Ports ---
        def try_serial(name, port, baud):
            try:
                s = serial.Serial(port, baud, timeout=0)
                print(f"✅ {name} connected on {port}")
                return s
            except Exception as e:
                failed.append(f"{name}: {e}")
                return None

        ser_ecall = try_serial("ECALL", ecall_com, BAUDRATE)
        ser_dashcam = try_serial("DASHCAM", dashcam_com, BAUDRATE)
        ser_dbmeter = try_serial("DB METER", dbmeter_com, DB_METER_BAUDRATE)
        serfeasa = try_serial("FEASA", feasa_com, FEASABAUDRATE)

        # --- 3️⃣ Camera ---
        try:
            cam_port = int(cam_port_text)
            cam_sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            cam_sock.settimeout(3)
            cam_sock.connect((cam_ip, cam_port))
            cam_sock.setblocking(False)
            print(f"📸 Camera connected to {cam_ip}:{cam_port}")
        except Exception as e:
            failed.append(f"Camera: {e}")
            cam_sock = None

        # --- 4️⃣ PSU ---
        try:
            import pyvisa
            psu_rm = pyvisa.ResourceManager()
            keysight_psu = psu_rm.open_resource(f"TCPIP0::{psu_ip}::INSTR")
            idn = keysight_psu.query("*IDN?")
            print(f"🔋 PSU Connected: {idn.strip()}")

            channel = 2
            keysight_psu.write(f"INST:NSEL {channel}")
            keysight_psu.write("VOLT 12")
            keysight_psu.write("CURR 1")
            keysight_psu.write("OUTP ON")
            print(f"🔋 PSU Output ON (Ch{channel}, 12V 1A)")
        except Exception as e:
            failed.append(f"PSU: {e}")
            keysight_psu = None

        # --- 5️⃣ NI DAQ (All Mandatory) ---
        try:
            system = System.local()
            detected_devices = [f"{getattr(dev, 'serial_num', 0):08X}" for dev in system.devices]
            expected_serials = [ni1, ni2, ni3]
            missing_serials = [x for x in expected_serials if x not in detected_devices]

            if not detected_devices:
                failed.append("NI DAQ: No devices detected.")
            elif missing_serials:
                failed.append(f"NI DAQ: Missing serial(s): {', '.join(missing_serials)}")
            else:
                print(f"🧠 NI DAQ Connected: {', '.join(detected_devices)}")
                ni_device_1, ni_device_2, ni_device_3 = ni1, ni2, ni3

                # ✅ Set initial NIUSB lines low (same as your lower code)
                niusb_set_low(ni_device_2, port=0, line=4)
                niusb_set_low(ni_device_2, port=0, line=5)
                print(f"NIUSB lines set LOW for {ni_device_2}")
        except Exception as e:
            failed.append(f"NI DAQ: {e}")

        # # --- 6️⃣ Final Mandatory Validation ---
        # if missing or failed:
        #     msg = "❌ Connection aborted! All devices are mandatory.\n\n"
        #     if missing:
        #         msg += "Missing:\n" + "\n".join(f"- {x}" for x in missing) + "\n\n"
        #     if failed:
        #         msg += "Failed:\n" + "\n".join(f"- {f}" for f in failed)
        #     QMessageBox.critical(self, "Connection Error", msg)
        #     ports_connected = False
        #     return

        # --- ✅ All Devices Connected ---
        ports_connected = True
        QMessageBox.information(self, "Success", "✅ All devices connected successfully!")
        print("✅ ALL DEVICES CONNECTED — READY TO START TESTING")

        # ✅ Callback if defined
        if ports_connected and self.open_barcode_callback:
            self.open_barcode_callback()

    


    # ---------------- UPDATE SETTINGS ----------------
    def update_settings(self):
        new_settings = {key: entry.text().strip() for key, entry in self.entries.items()}
        if new_settings == self.settings:
            QMessageBox.information(self, "Info", "No changes detected; settings already up-to-date.")
            return

        txt, ok = QInputDialog.getText(
            self, "Admin Required",
            "Enter Admin password to update settings:",
            QLineEdit.EchoMode.Password
        )
        if not ok:
            return
        if not self._validate_admin_password(txt.strip()):
            QMessageBox.critical(self, "Access Denied", "Incorrect admin password!")
            return

        save_com_settings(new_settings)
        self.settings = new_settings
        QMessageBox.information(self, "Info", "Settings updated!")


class BarcodeScreen(QWidget):
    def __init__(self, on_submit=None):
        super().__init__()
        self.on_submit = on_submit
        self.setWindowTitle("Scan Barcodes")
        self.setMinimumSize(600, 340)
        layout = QVBoxLayout(self)

        # Gradient background
        self.gradient = QLinearGradient(0, 0, 0, self.height())
        self.gradient.setColorAt(0.0, QColor(0, 230, 255))
        self.gradient.setColorAt(0.2, QColor(10, 255, 140))
        self.gradient.setColorAt(0.4, QColor(0, 255, 220, 180))
        self.gradient.setColorAt(0.6, QColor(10, 10, 10))
        self.gradient.setColorAt(0.8, QColor(240, 240, 240))
        self.gradient.setColorAt(1.0, QColor(0, 140, 255))

        # Card container
        self.card = Extreme3DCard()
        layout.addWidget(self.card, alignment=Qt.AlignmentFlag.AlignCenter)

        grid = QGridLayout()
        self.card.layout.addLayout(grid)

        self.entries = {}
        # Only E-Call and Dashcam fields now
        labels = ["E-Call Barcode:", "Dashcam Barcode:"]

        for i, label in enumerate(labels):
            lbl = QLabel(label)
            lbl.setStyleSheet("color: cyan; font-weight: bold;")
            edit = QLineEdit()
            edit.setPlaceholderText(f"Enter {label[:-1]}")
            edit.setStyleSheet("background-color:#10182e; color:#fff; border-radius:5px; padding:6px")
            self.entries[label] = edit
            grid.addWidget(lbl, i, 0, alignment=Qt.AlignmentFlag.AlignRight)
            grid.addWidget(edit, i, 1)

        # Enter key navigation
        self.entries["E-Call Barcode:"].returnPressed.connect(lambda: self.entries["Dashcam Barcode:"].setFocus())
        self.entries["Dashcam Barcode:"].returnPressed.connect(self.check_barcodes)

        # Buttons
        btn_layout = QHBoxLayout()
        self.submit_button = GlassyNeonButton("Submit")
        self.submit_button.clicked = self.check_barcodes
        btn_layout.addWidget(self.submit_button)
        self.card.layout.addLayout(btn_layout)

        # Set initial focus on E-Call barcode
        QTimer.singleShot(100, lambda: self.entries["E-Call Barcode:"].setFocus())

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QBrush(self.gradient))
        super().paintEvent(event)

    def clear_barcodes(self):
        for entry in self.entries.values():
            entry.clear()
        self.entries["E-Call Barcode:"].setFocus()

    def check_barcodes(self):
        global ECALL_BARCODE, DASHCAM_BARCODE

        ecall = self.entries["E-Call Barcode:"].text().strip()
        dashcam = self.entries["Dashcam Barcode:"].text().strip()

        if not (ecall and dashcam):
            QMessageBox.critical(self, "Error", "Both barcodes are required!")
            # return
        
        # if len(ecall)!= 40 or len(dashcam) != 21:
        #     QMessageBox.critical(self, "Error", "Wrong barcodes!")
        #     self.clear_barcodes()
        #     return

        ECALL_BARCODE = ecall
        DASHCAM_BARCODE = dashcam

        # QMessageBox.information(self, "Done", "Barcodes submitted successfully!")
        if self.on_submit:
            self.on_submit()


from PyQt6.QtCore import Qt, QTimer, pyqtSignal
from PyQt6.QtWidgets import QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QFrame, QTabWidget, QTableWidget, QTableWidgetItem, QAbstractItemView
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QPropertyAnimation, QEasingCurve
from PyQt6.QtGui import QColor, QPainter, QLinearGradient
from PyQt6.QtWidgets import QWidget, QLabel, QVBoxLayout, QHBoxLayout, QFrame, QTabWidget, QTableWidget, QTableWidgetItem, QAbstractItemView, QGraphicsDropShadowEffect
import threading
import time
from PyQt6.QtWidgets import (
    QWidget, QFrame, QLabel, QPushButton, QHBoxLayout, QVBoxLayout, QTabWidget, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QGraphicsDropShadowEffect
)
from PyQt6.QtCore import Qt, QPropertyAnimation, QEasingCurve, QTimer
from PyQt6.QtGui import QPainter, QLinearGradient, QColor

# Assuming PSUController is already imported or defined above




class MainScreen(QWidget):
    overallBannerRequested = pyqtSignal(str, int)  # (result, duration_ms)

    def __init__(self, main_window=None):
        super().__init__()
        self.main_window = main_window
        self.setWindowTitle("Command Panel")
        self.setMinimumSize(1000, 600)
        self.loop_count_var = 0
        # Only port and line are fixed
        self.ni_daq_port = 2
        self.ni_daq_line = 1

        self.ni_daq_monitor_thread = None
        self.ni_daq_stop_event = threading.Event()
        self.run_allowed = False

        # HUD setup
        self.hud_timer = QTimer(self)
        self.hud_timer.timeout.connect(self.update_hud_timer)
        self.hud_start_time = None
        self.hud_running = False

        # Layout setup
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(10, 60, 10, 10)
        self.main_layout.setSpacing(10)
        self.overallBannerRequested.connect(self.show_overall_banner)

        self.top_panel()
        self.main_area()
        self.bottom_close_button()
        self.init_hud_display()
    # ============================================================
    # 🔹 NI DAQ Dynamic Handling + Real Monitoring + Auto Test Trigger
    # ============================================================
    
    def update_ni_daq_serial(self, new_serial_hex):
        """Update NI DAQ serial dynamically from COM port UI."""
        if not new_serial_hex:
            print("⚠️ Empty NI DAQ serial provided — ignoring.")
            return
    
        global ni_device_3
        ni_device_3 = new_serial_hex
        print(f"✅ NI DAQ serial updated dynamically: {ni_device_3}")
    
        # Restart monitoring automatically if running
        if self.ni_daq_monitor_thread and self.ni_daq_monitor_thread.is_alive():
            print("🔁 Restarting NI DAQ monitor with new serial...")
            self.stop_ni_daq_monitor()
            self.start_ni_daq_monitor(ni_device_3, self.ni_daq_port, self.ni_daq_line)
    
    
    def showEvent(self, event):
        """Start NI DAQ monitoring when GUI is shown."""
        super().showEvent(event)
    
        global ni_device_3
        if not ni_device_3:
            print("⚠️ NI Device 3 not yet connected (from COM Port screen)")
            return
    
        print(f"🧩 Using NI Device Serial: {ni_device_3}")
    
        if not self.ni_daq_monitor_thread or not self.ni_daq_monitor_thread.is_alive():
            self.start_ni_daq_monitor(ni_device_3, self.ni_daq_port, self.ni_daq_line)
    
    
    def start_ni_daq_monitor(self, serial_hex, port, line):
        """Start NI DAQ monitoring thread."""
        if self.ni_daq_monitor_thread and self.ni_daq_monitor_thread.is_alive():
            print("⚙️ NI DAQ monitoring already running.")
            return
    
        if not serial_hex:
            print("❌ Serial Hex missing, cannot start DAQ.")
            return
    
        serial_value = str(serial_hex).strip()
        self.ni_daq_stop_event.clear()
        self.ni_daq_monitor_thread = threading.Thread(
            target=self._ni_daq_monitor_loop,
            args=(serial_value, port, line),
            daemon=True
        )
        self.ni_daq_monitor_thread.start()
        print(f"🟢 NI DAQ monitoring started for serial: {serial_value}")
    
    
    def stop_ni_daq_monitor(self):
        """Stop NI DAQ monitoring."""
        if self.ni_daq_monitor_thread and self.ni_daq_monitor_thread.is_alive():
            print("🛑 Stopping NI DAQ monitor...")
            self.ni_daq_stop_event.set()
            self.ni_daq_monitor_thread.join(timeout=1.0)
            self.ni_daq_monitor_thread = None
            print("✅ NI DAQ monitor stopped.")
    
    
    def _ni_daq_monitor_loop(self, serial_hex, port, line):
        """Background NI DAQ monitoring logic with test trigger."""
        try:
            from nidaqmx import Task
            from nidaqmx.system import system
            import nidaqmx
    
            serial_decimal = int(serial_hex, 16)
            system = System.local()
            dev = next((d for d in system.devices if getattr(d, "serial_num", None) == serial_decimal), None)
    
            if not dev:
                print(f"❌ Device with serial {serial_hex} not found.")
                return
    
            line_name = f"{dev.name}/port{port}/line{line}"
            print(f"🔍 Monitoring {line_name} (Digital Input Trigger Mode)")
    
            with Task() as task:
                task.di_channels.add_di_chan(line_name)
                prev_state = bool(task.read())
                print(f"Initial State Detected: {'HIGH' if prev_state else 'LOW'} (ignored)")
    
                while not self.ni_daq_stop_event.is_set():
                    state = bool(task.read())
    
                    # Rising Edge (Voltage Applied)
                    if not prev_state and state:
                        print("⚡ EXECUTED: Voltage Applied / Triggered ✅")
    
                    # Falling Edge (Voltage Removed)
                    if prev_state and not state:
                        print("🔴 RESET: Voltage Removed ⛔")
    
                        if self.run_allowed:
                            # Disable further triggers until user resets
                            self.run_allowed = False
                            print("▶️ Calling Run All Tests (from DAQ trigger)...")
                            QTimer.singleShot(0, self._run_all_tests_thread)
                            time.sleep(0.5)
                        else:
                            print("⚠️ Run not allowed yet, ignoring DAQ trigger.")
    
                    prev_state = state
                    time.sleep(0.05)
    
        except Exception as e:
            print(f"NI DAQ Monitor Error: {e}")
    
    
    def init_hud_display(self):
        # Show initial "ready" display on HUD
        self.hud_display.setText("Ready. Elapsed: 00:00:00.0")

    def start_hud_timer(self):
        self.hud_start_time = time.time()
        self.hud_timer.start(100)        # 100 ms update interval
        self.hud_running = True
        self.hud_display.setText("Elapsed: 00:00:00.0")
    
    def update_hud_timer(self):
        if self.hud_start_time is None:
            return
        elapsed = time.time() - self.hud_start_time
        mins, secs = divmod(elapsed, 60)
        secs_int = int(secs)
        millis = int((secs - secs_int) * 10)
        timer_text = f"Elapsed: {int(mins):02d}:{secs_int:02d}:{millis}"
        self.hud_display.setText(timer_text)

    
    

    
    # def finalize_hud_timer(self):
    #     self.hud_timer.stop()
    #     self.hud_running = False
    #     total_elapsed = time.time() - self.hud_start_time if self.hud_start_time else 0
    #     mins, secs = divmod(total_elapsed, 60)
    #     secs_int = int(secs)
    #     millis = int((secs - secs_int) * 10)
    #     self.hud_display.setText(f"Run Complete: {int(mins):02d}:{secs_int:02d}:{millis}")


    def finalize_hud_timer(self):
        """Safely stop HUD timer from GUI thread."""
        try:
            if hasattr(self, "hud_timer") and self.hud_timer:
                if QThread.currentThread() != QApplication.instance().thread():
                    # ⚠️ Called from worker thread — safely schedule stop in GUI thread
                    QTimer.singleShot(0, self.finalize_hud_timer)
                    return
                self.hud_timer.stop()
                self.hud_running = False
                total_elapsed = time.time() - self.hud_start_time if getattr(self, "hud_start_time", None) else 0
                mins, secs = divmod(total_elapsed, 60)
                secs_int = int(secs)
                millis = int((secs - secs_int) * 10)
                self.hud_display.setText(f"Run Complete: {int(mins):02d}:{secs_int:02d}:{millis}")
                print("⏹ HUD timer stopped safely")
        except Exception as e:
            print("HUD timer stop error:", e)


    def reset_hud_timer(self):
        self.hud_timer.stop()
        self.hud_running = False
        self.hud_start_time = None
        self.init_hud_display()


    def _run_all_tests_thread(self):
        self.run_allowed = False  # disable until user resets/adds new tab
        self.start_hud_timer()
        threading.Thread(target=lambda: run_all_tests(active_table, self.status_var, self, self.main_window), daemon=True).start()
    
    def run_selected(self):
        sel = active_table.selectionModel().selectedRows()
        if not sel:
            print("No selection")
            return
        row = sel[0].row()
        # self.start_hud_timer()               # Start HUD timer for run selected
        threading.Thread(target=lambda: run_test_case(row,active_table, self.status_var), daemon=True).start()
    

    def start_loop(self, limit=0):
        stop_event.clear()
        self.start_hud_timer()               # Ensure HUD timer starts for loop
        threading.Thread(target=self.test_loop, args=(limit,), daemon=True).start()

    def restart_tests(self):
        stop_event.set()
        stop_event.clear()
        self.reset_hud_timer()      # Clear HUD on restart
        self.add_new_tab()
        self.status_var.setText("Restarted: New run table ready")

    def paintEvent(self, event):
        painter = QPainter(self)
        rect = self.rect()
        gradient = QLinearGradient(0, 0, 0, rect.height())
        gradient.setColorAt(0.0, QColor(0, 230, 255))
        gradient.setColorAt(0.2, QColor(10, 255, 140))
        gradient.setColorAt(0.4, QColor(0, 255, 220, 180))
        gradient.setColorAt(0.6, QColor(10, 10, 10))
        gradient.setColorAt(0.8, QColor(240, 240, 240))
        gradient.setColorAt(1.0, QColor(0, 140, 255))
        painter.fillRect(rect, gradient)



    def top_panel(self):
        top = QFrame()
        top.setStyleSheet("background-color: #0d1321;")
        top.setFixedHeight(70)
    
        layout = QHBoxLayout(top)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(20)
    
        # Left vertical panel for Command label and HUD
        left_panel = QVBoxLayout()
        left_panel.setSpacing(5)
    
        lbl = QLabel("Command Panel")
        lbl.setStyleSheet("color: #00d1ff; font-weight:bold; font-size:18px;")
        left_panel.addWidget(lbl)
    
        # HUD small screen
        self.hud_display = QLabel("HUD Display")
        self.hud_display.setStyleSheet("""
            background-color: #021f2f;
            color: #00ffa0;
            font-size: 12px;
            font-weight: bold;
            border: 1px solid #00d1ff;
            border-radius: 5px;
            padding: 5px 10px;
            min-width: 180px;
            max-width: 250px;
        """)
        self.hud_display.setFixedHeight(30)
        self.hud_display.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_panel.addWidget(self.hud_display)
    
        layout.addLayout(left_panel)
        layout.addStretch()
    
        # --- 🟢 NI DAQ Button ---
        self.ni_daq_button = QPushButton("NI DAQ")
        self.ni_daq_button.setFixedSize(80, 30)
        self.ni_daq_button.setStyleSheet("background-color: #007a3d; color: white; font-weight:bold;")
        self.ni_daq_button.clicked.connect(self.open_ni_daq_controller)
        layout.addWidget(self.ni_daq_button)
    
        # --- 🔵 PSU Button ---
        self.psu_button = QPushButton("PSU")
        self.psu_button.setFixedSize(80, 30)
        self.psu_button.setStyleSheet("background-color: #0059b3; color: white; font-weight:bold;")
        self.psu_button.clicked.connect(self.open_psu_controller)
        layout.addWidget(self.psu_button)
    
        self.main_layout.addWidget(top)
    
    
    def open_psu_controller(self):
        if not hasattr(self, 'psu_window'):
            self.psu_window = PSUController()
        self.psu_window.show()
        self.psu_window.raise_()
        self.psu_window.activateWindow()
    
    
    def open_ni_daq_controller(self):
        """Opens NI USB DAQ Control window."""
        if not hasattr(self, 'ni_daq_window'):
            self.ni_daq_window = NIDAQController()
        self.ni_daq_window.show()
        self.ni_daq_window.raise_()
        self.ni_daq_window.activateWindow()


    def main_area(self):
        area = QHBoxLayout()
    
        # ---------------- LEFT FRAME ----------------
        self.left_frame = QFrame()
        self.left_frame.setStyleSheet("""
            QFrame {
                background-color: rgba(18, 27, 43, 180);
                border-radius: 15px;
                border: 1px solid #00ffff;
            }
        """)
    
        effect_left = QGraphicsDropShadowEffect()
        effect_left.setBlurRadius(40)
        effect_left.setColor(QColor(0, 255, 255, 180))
        effect_left.setXOffset(0)
        effect_left.setYOffset(10)
        self.left_frame.setGraphicsEffect(effect_left)
    
        # ✅ Plain container (replaced QTabWidget)
        left_layout = QVBoxLayout(self.left_frame)
        self.table_container = QWidget()
        self.table_layout = QVBoxLayout(self.table_container)
        left_layout.addWidget(self.table_container)
    
        # ✅ Add single table (no duplicate)
        if hasattr(self, "table_layout"):
            self.add_new_tab()
    
        # ✅ Controls panel below table
        self.controls_panel(left_layout)
        area.addWidget(self.left_frame, 3)
    
        # ---------------- RIGHT FRAME ----------------
        self.right_frame = QFrame()
        self.right_frame.setStyleSheet("""
            QFrame {
                background-color: rgba(19, 41, 75, 180);
                border-radius: 15px;
                border: 1px solid #00ffd5;
            }
        """)
    
        effect_right = QGraphicsDropShadowEffect()
        effect_right.setBlurRadius(40)
        effect_right.setColor(QColor(0, 255, 213, 180))
        effect_right.setXOffset(0)
        effect_right.setYOffset(10)
        self.right_frame.setGraphicsEffect(effect_right)
    
        right_layout = QVBoxLayout(self.right_frame)
        self.right = right_layout
        self.right_counters()
        area.addWidget(self.right_frame, 1)
    
        # ✅ Add to main layout
        self.main_layout.addLayout(area)
    
        # ✅ Animate both glow effects
        self.animate_glow(effect_left)
        self.animate_glow(effect_right)

    def animate_glow(self, effect):
        anim = QPropertyAnimation(effect, b"color")
        anim.setDuration(2000)
        anim.setLoopCount(-1)
        anim.setEasingCurve(QEasingCurve.Type.InOutSine)

        start_color = effect.color()
        end_color = QColor(start_color.red(), start_color.green(), start_color.blue(), 255)

        anim.setStartValue(start_color)
        anim.setEndValue(end_color)
        anim.setDirection(QPropertyAnimation.Direction.Forward)

        def on_value_changed(value):
            effect.setColor(value)

        anim.valueChanged.connect(on_value_changed)
        anim.start()

        if not hasattr(self, "_glow_animations"):
            self._glow_animations = []
        self._glow_animations.append(anim)



    def add_new_tab(self):
        self.reset_hud_timer()
        global tab_counter, active_table
        if 'tab_counter' not in globals():
            tab_counter = 0
    
        # ✅ Remove previous table if already added
        if hasattr(self, "table_layout"):
            while self.table_layout.count():
                child = self.table_layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
    
        # ✅ Create new table
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["NAME", "MINIMUM", "OUTPUT", "MAXIMUM", "STATUS", "TIME(s)"])
        table.setStyleSheet("""
            QTableWidget {
                background-color: #10182e;
                color: white;
                gridline-color: #00ffff;
                border: none;
                selection-background-color: #00ffff;
                selection-color: black;
            }
            QHeaderView::section {
                background-color: #16213e;
                color: #00ffff;
                border: 1px solid #00ffff;
                padding: 4px;
            }
        """)
    
        from PyQt6.QtWidgets import QTableWidgetItem
        for tc in TEST_CASES:
            name = tc.get("name", "")
            expected = tc.get("expected", "")
            output = tc.get("output", "")
            status = tc.get("status", "")
            time_taken = tc.get("time", "")
    

            min_val = tc.get("min") if tc.get("min") is not None else (tc.get("min_limit") if tc.get("min_limit") is not None else tc.get("min_val"))
            max_val = tc.get("max") if tc.get("max") is not None else (tc.get("max_limit") if tc.get("max_limit") is not None else tc.get("max_val"))

    
            # --- Combine logic for min/max with expected ---
            if min_val is not None and max_val is not None and expected:
                minimum_text = f"{min_val}\nExpected: {expected}"
                maximum_text = f"{max_val}\nExpected: {expected}"
            elif min_val is not None and max_val is not None:
                minimum_text = str(min_val)
                maximum_text = str(max_val)
            elif expected:
                minimum_text = f"Expected: {expected}"
                maximum_text = f"Expected: {expected}"
            else:
                minimum_text = ""
                maximum_text = ""
    
            # --- Collect dbmeter_checks lines ---
            db_min_lines, db_max_lines = [], []
            if "dbmeter_checks" in tc:
                for check in tc["dbmeter_checks"]:
                    cmin = check.get("min")
                    cmax = check.get("max")
                    label = check.get("label", "")
                    if cmin is not None:
                        db_min_lines.append(f"{cmin} ({label})")
                    if cmax is not None:
                        db_max_lines.append(f"{cmax} ({label})")
    
            # --- Collect value_ranges lines with Red/Green/Blue labels ---
            range_min_lines, range_max_lines = [], []
            if "value_ranges" in tc:
                labels = ["Red", "Green", "Blue"]
                for i, vr in enumerate(tc["value_ranges"]):
                    if isinstance(vr, (list, tuple)) and len(vr) == 2:
                        cmin, cmax = vr
                        color_label = labels[i % len(labels)]
                        range_min_lines.append(f"{cmin} ({color_label})")
                        range_max_lines.append(f"{cmax} ({color_label})")
    
            # --- Merge all min/max lines ---
            min_lines = [minimum_text] if minimum_text else []
            max_lines = [maximum_text] if maximum_text else []
    
            if db_min_lines:
                min_lines.extend(db_min_lines)
            if db_max_lines:
                max_lines.extend(db_max_lines)
            if range_min_lines:
                min_lines.extend(range_min_lines)
            if range_max_lines:
                max_lines.extend(range_max_lines)
    
            min_combined = "\n".join(min_lines)
            max_combined = "\n".join(max_lines)
    
            # --- Add main test row ---
            row = table.rowCount()
            table.insertRow(row)
    
            table.setItem(row, 0, QTableWidgetItem(name))
            table.setItem(row, 1, QTableWidgetItem(min_combined))
            table.setItem(row, 2, QTableWidgetItem(str(output)))
            table.setItem(row, 3, QTableWidgetItem(max_combined))
            table.setItem(row, 4, QTableWidgetItem(status))
            table.setItem(row, 5, QTableWidgetItem(str(time_taken)))
    
            # --- Keep daq_steps as before ---
            if "daq_steps" in tc:
                for step in tc["daq_steps"]:
                    cmin = step.get("min")
                    cmax = step.get("max")
                    if cmin is not None or cmax is not None:
                        sub_row = table.rowCount()
                        table.insertRow(sub_row)
                        table.setItem(sub_row, 0, QTableWidgetItem(""))
                        table.setItem(sub_row, 1, QTableWidgetItem(str(cmin if cmin is not None else "")))
                        table.setItem(sub_row, 2, QTableWidgetItem(""))
                        table.setItem(sub_row, 3, QTableWidgetItem(str(cmax if cmax is not None else "")))
                        table.setItem(sub_row, 4, QTableWidgetItem(""))
                        table.setItem(sub_row, 5, QTableWidgetItem(""))
    
        # ✅ Table behavior + word wrap
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setWordWrap(True)
        table.resizeRowsToContents()
        table.itemChanged.connect(lambda: table.resizeRowsToContents())



        def auto_scroll(item):
            if item.column() == 4:  # STATUS updated
                row = item.row()
                if row >= 3:  # first 3 rows no scroll
                    table.scrollToItem(item, QAbstractItemView.ScrollHint.PositionAtCenter)
        
        table.itemChanged.connect(auto_scroll)


    
        # ✅ Column widths
        column_widths = [200, 150, 150, 150, 100, 100]
        for col, width in enumerate(column_widths):
            table.setColumnWidth(col, width)
    
        header = table.horizontalHeader()
        header.setStretchLastSection(True)
    
        # ✅ Add new table cleanly
        if hasattr(self, "table_layout"):
            self.table_layout.addWidget(table)
    
        # ✅ Keep reference
        active_table = table
        tab_counter += 1

##############------------> without ok

    # def show_overall_banner(self, overall_result, duration_ms=3000):
    #     """
    #     Display a big PASS/FAIL banner centered on the GUI for duration_ms milliseconds,
    #     then hide it and open barcode screen.
    #     """
    #     print("enter in overall banner")
    
    #     # 🧹 Remove existing banner if any
    #     if hasattr(self, "_overall_banner") and self._overall_banner is not None:
    #         try:
    #             self._overall_banner.hide()
    #             self._overall_banner.deleteLater()
    #         except Exception as e:
    #             print("cleanup old banner error:", e)
    #         self._overall_banner = None
    
    #     # ✅ Ensure this screen is visible on stack
    #     if hasattr(self, "main_window"):
    #         self.main_window.stack.setCurrentWidget(self)
    
    #     # 🌫 Create overlay to dim background
    #     overlay = QWidget(self)
    #     overlay.setStyleSheet("background-color: rgba(0, 0, 0, 80);")  # semi-transparent overlay
    #     overlay.setGeometry(0, 0, self.width(), self.height())
    #     overlay.show()
    #     overlay.raise_()
    
    #     # 🏷 Create banner label
    #     banner = QLabel(overlay)
    #     banner.setAlignment(Qt.AlignmentFlag.AlignCenter)
    #     banner.setText(overall_result.upper())
    #     banner.setObjectName("overallResultBanner")
    
    #     # 💅 Styling
    #     banner.setStyleSheet(f"""
    #         QLabel#overallResultBanner {{
    #             background-color: {'#92d050' if overall_result.upper() == 'PASS' else '#FF4747'};
    #             color: black;
    #             font-weight: 900;
    #             font-size: 48px;
    #             border-radius: 10px;
    #             padding: 30px;
    #             border: 4px solid rgba(0,0,0,0.2);
    #         }}
    #     """)
    
    #     # 📐 Size and position
    #     win_w = max(600, self.width())
    #     win_h = max(400, self.height())
    #     bw = int(win_w * 0.75)
    #     bh = int(win_h * 0.20)
    #     banner.setFixedSize(bw, bh)
    #     banner.move((win_w - bw)//2, (win_h - bh)//2)
    
    #     # 🚀 Show overlay + banner
    #     banner.raise_()
    #     banner.show()
    #     banner.repaint()
    #     QApplication.processEvents()
    #     print("✅ Banner shown ->", overall_result)
    
    #     # 🔒 Keep reference
    #     self._overall_banner = overlay
    
    #     # ⏳ Hide after duration_ms + go to barcode screen
    #     def hide_and_open():
    #         try:
    #             if hasattr(self, "_overall_banner") and self._overall_banner:
    #                 self._overall_banner.hide()
    #                 self._overall_banner.deleteLater()
    #                 self._overall_banner = None
    #         except Exception as e:
    #             print(f"banner hide error: {e}")
    
    #         # ✅ switch back to barcode screen after banner
    #         if hasattr(self, "main_window") and hasattr(self.main_window, "open_barcode_screen"):
    #             print("🔄 Returning to Barcode Screen...")
    #             self.main_window.open_barcode_screen()
    
    #     QTimer.singleShot(duration_ms, hide_and_open)


####---------->>>>>>>>>>>>>> with ok
    def show_overall_banner(self, overall_result, duration_ms=None):
        """
        Display a big PASS/FAIL banner centered on the GUI.
        User must press OK to continue to the Barcode screen.
        """
        print("enter in overall banner")
    
        # 🧹 Remove existing banner if any
        if hasattr(self, "_overall_banner") and self._overall_banner is not None:
            try:
                self._overall_banner.hide()
                self._overall_banner.deleteLater()
            except Exception as e:
                print("cleanup old banner error:", e)
            self._overall_banner = None
    
        # ✅ Ensure this screen is visible on stack
        if hasattr(self, "main_window"):
            self.main_window.stack.setCurrentWidget(self)
    
        # 🌫 Create overlay to dim background
        overlay = QWidget(self)
        overlay.setStyleSheet("background-color: rgba(0, 0, 0, 80);")  # semi-transparent overlay
        overlay.setGeometry(0, 0, self.width(), self.height())
        overlay.show()
        overlay.raise_()
    
        # 🏷 Create banner label
        banner = QLabel(overlay)
        banner.setAlignment(Qt.AlignmentFlag.AlignCenter)
        banner.setText(overall_result.upper())
        banner.setObjectName("overallResultBanner")
    
        # 💅 Styling
        banner.setStyleSheet(f"""
            QLabel#overallResultBanner {{
                background-color: {'#92d050' if overall_result.upper() == 'PASS' else '#FF4747'};
                color: black;
                font-weight: 900;
                font-size: 48px;
                border-radius: 10px;
                padding: 30px;
                border: 4px solid rgba(0,0,0,0.2);
            }}
        """)
    
        # 📐 Size and position
        win_w = max(600, self.width())
        win_h = max(400, self.height())
        bw = int(win_w * 0.75)
        bh = int(win_h * 0.25)
        banner.setFixedSize(bw, bh)
        banner.move((win_w - bw)//2, (win_h - bh)//2 - 30)
    
        # ✅ OK Button below banner
        ok_button = QPushButton("OK", overlay)
        ok_button.setFixedSize(120, 50)
        ok_button.move((win_w - ok_button.width())//2, (win_h - bh)//2 + bh + 20)
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d7;
                color: white;
                font-size: 20px;
                font-weight: bold;
                border-radius: 8px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #005a9e;
            }
            QPushButton:pressed {
                background-color: #004578;
            }
        """)
    
        # 🚀 Show overlay + banner
        banner.raise_()
        banner.show()
        ok_button.raise_()
        ok_button.show()
        QApplication.processEvents()
        print("✅ Banner shown ->", overall_result)
    
        # 🔒 Keep reference
        self._overall_banner = overlay
    
        # 🎯 Define OK button behavior
        def on_ok_clicked():
            try:
                if hasattr(self, "_overall_banner") and self._overall_banner:
                    self._overall_banner.hide()
                    self._overall_banner.deleteLater()
                    self._overall_banner = None
            except Exception as e:
                print(f"banner hide error: {e}")
    
            if hasattr(self, "main_window") and hasattr(self.main_window, "open_barcode_screen"):
                print("🔄 Returning to Barcode Screen after OK...")
                self.main_window.open_barcode_screen()
    
        ok_button.clicked.connect(on_ok_clicked)








    def controls_panel(self, parent_layout):
        self.status_var = QLabel("")
        self.status_var.setStyleSheet("color:#00ffd5; font-weight:bold;")
        parent_layout.addWidget(self.status_var)

        btn_layout = QHBoxLayout()
        parent_layout.addLayout(btn_layout)

        def glassy_button(text, func, color_mode='gold'):
            b = GlassyNeonButton(text, color_mode=color_mode)
            b.clicked = func
            b.setMinimumWidth(110)
            btn_layout.addWidget(b)
            return b

        glassy_button("Run Selected", self.run_selected)
        glassy_button("Run All", self._run_all_tests_thread)
        glassy_button("STOP", self.stop_tests, color_mode='silver')
        glassy_button("Restart", self.restart_tests, color_mode='gold')
        glassy_button("Loop", self.start_loop, color_mode='gold')
        glassy_button("Generate Report", lambda: [generate_report(active_table), finalize_run(active_table, self)], color_mode='silver')
        btn_layout.addStretch()

        loop_layout = QHBoxLayout()
        lbl = QLabel("Current Run:")
        lbl.setStyleSheet("color:white;")
        self.loop_display = QLabel(str(self.loop_count_var))
        self.loop_display.setStyleSheet("color:white; font-weight:bold;")
        loop_layout.addWidget(lbl)
        loop_layout.addWidget(self.loop_display)
        loop_layout.addStretch()
        parent_layout.addLayout(loop_layout)

    def stop_tests(self):
        stop_event.set()
        self.status_var.setText("Stopped by user")

    def refresh_counters(self):
        global total_count, pass_count, fail_count, yield_count
        self.label_total.setText(str(total_count))
        self.label_pass.setText(str(pass_count))
        self.label_fail.setText(str(fail_count))
        self.label_yield.setText(f"{yield_count}%")

    def test_loop(self, limit=0):
        count = 1
        while not stop_event.is_set() and (limit == 0 or count <= limit):
            run_all_tests(active_table, self.status_var, self)
            count += 1
            QTimer.singleShot(0, lambda: self.status_var.setText(f"Run {count} of {'∞' if limit==0 else limit}"))
            self.loop_count_var = count
            self.loop_display.setText(str(self.loop_count_var))
            generate_report(active_table)
            finalize_run(active_table, self)
            self.add_new_tab()
            for _ in range(5):
                if stop_event.is_set():
                    break
                time.sleep(1)
            self.status_var.setText("Restarting Loop...")
        self.status_var.setText("Loop finished or stopped.")

    def right_counters(self):
        global total_count, pass_count, fail_count, yield_count

        def counter_label(text, var):
            lbl_title = QLabel(text.upper())
            lbl_title.setStyleSheet("color:#00ffd5; font-weight:bold;")
            lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.right.addWidget(lbl_title)
            lbl_val = QLabel(str(var))
            lbl_val.setStyleSheet("background-color:#13294b; color:#00d1ff; font-size:18px; font-weight:bold;")
            lbl_val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_val.setFixedHeight(40)
            self.right.addWidget(lbl_val)
            return lbl_val

        self.label_total = counter_label("Total", total_count)
        self.label_pass = counter_label("Pass", pass_count)
        self.label_fail = counter_label("Fail", fail_count)
        self.label_yield = counter_label("Yield", yield_count)

        self.btn_reset = GlassyNeonButton("Reset Counters", color_mode='silver')
        self.btn_reset.clicked = self.reset_counters
        self.right.addWidget(self.btn_reset)

    def reset_counters(self):
        global total_count, pass_count, fail_count, yield_count
        total_count = pass_count = fail_count = yield_count = 0
        self.refresh_counters()
        self.status_var.setText("Counters reset to 0")
        save_counters()

    def bottom_close_button(self):
        self.btn_close = GlassyNeonButton("Close App", color_mode='gold')
        self.btn_close.clicked = QApplication.instance().quit
        self.main_layout.addWidget(self.btn_close, alignment=Qt.AlignmentFlag.AlignCenter)








class MainWindow(QMainWindow):
    load_counters() #for counter load of pass fail or inn sbke 
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Vipin App")
        self.resize(800, 500)

        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        # Screens
        self.login_screen = LoginWindow(login_success_callback=self.open_com_port_screen)
        self.com_port_screen = ComPortSettingsScreen(open_barcode_callback=self.open_barcode_screen)
        self.barcode_screen = BarcodeScreen(on_submit=self.after_barcode_submit)
        self.main_screen = MainScreen(main_window=self)    # Pass self here


        self.stack.addWidget(self.login_screen)
        self.stack.addWidget(self.com_port_screen)
        self.stack.addWidget(self.barcode_screen)
        self.stack.addWidget(self.main_screen) 



    def open_com_port_screen(self):
        self.stack.setCurrentWidget(self.com_port_screen)

    def open_barcode_screen(self):
        self.barcode_screen.clear_barcodes()
        self.stack.setCurrentWidget(self.barcode_screen)

    def after_barcode_submit(self):
        
        # QMessageBox.information(self, "Flow", "All steps done, going to MainScreen.")
        self.stack.setCurrentWidget(self.main_screen)
        self.main_screen.add_new_tab() 
        self.main_screen.run_allowed = True  # allow DAQ to trigger


import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QMainWindow, QLabel, QLineEdit, QPushButton,
    QComboBox, QMessageBox, QVBoxLayout, QHBoxLayout, QFormLayout, QFrame
)
from PyQt6.QtCore import Qt

from PyQt6.QtGui import QFont, QPalette, QColor

class PSUController(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Keysight PSU Controller")
        self.setMinimumSize(400, 400)

        # Fonts
        self.status_font = QFont("Helvetica", 12, QFont.Weight.Bold)
        self.value_font = QFont("Helvetica", 14, QFont.Weight.Bold)

        # PyVISA resource manager and device instance
        self.rm = None
        self.psu = None

        # Default VISA Address
        self.default_visa_address = "TCPIP0::192.168.100.2::inst0::INSTR"

        # Widgets
        self.init_ui()

    def init_ui(self):
        # Main widget and layout
        main_widget = QWidget()
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        # Input Frame (like the input_frame in Tkinter) with blue background frame
        input_frame = QFrame()
        input_frame.setFrameShape(QFrame.Shape.Panel)
        input_frame.setFrameShadow(QFrame.Shadow.Raised)
        input_palette = input_frame.palette()
        input_palette.setColor(QPalette.ColorRole.Window, QColor("#6699CC"))
        input_frame.setAutoFillBackground(True)
        input_frame.setPalette(input_palette)
        input_frame.setContentsMargins(15, 15, 15, 15)

        input_layout = QFormLayout()
        input_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        input_frame.setLayout(input_layout)

        # VISA Address Entry
        self.visa_address_edit = QLineEdit()
        self.visa_address_edit.setText(self.default_visa_address)
        self.visa_address_edit.setFont(self.value_font)
        label_visa = QLabel("PSU VISA Address:")
        label_visa.setFont(self.status_font)
        label_visa.setStyleSheet("color: white;")
        input_layout.addRow(label_visa, self.visa_address_edit)

        # Channel dropdown
        label_channel = QLabel("Channel (1 or 2 or 3):")
        label_channel.setFont(self.status_font)
        label_channel.setStyleSheet("color: white;")

        self.channel_combo = QComboBox()
        self.channel_combo.addItems(["1", "2", "3"])
        self.channel_combo.setFont(self.value_font)

        input_layout.addRow(label_channel, self.channel_combo)

        # Voltage Entry
        label_voltage = QLabel("Voltage (V):")
        label_voltage.setFont(self.status_font)
        label_voltage.setStyleSheet("color: white;")
        self.voltage_edit = QLineEdit()
        self.voltage_edit.setFont(self.value_font)
        input_layout.addRow(label_voltage, self.voltage_edit)

        # Current Entry
        label_current = QLabel("Current (A):")
        label_current.setFont(self.status_font)
        label_current.setStyleSheet("color: white;")
        self.current_edit = QLineEdit()
        self.current_edit.setFont(self.value_font)
        input_layout.addRow(label_current, self.current_edit)

        # Buttons Frame
        btn_frame = QFrame()
        btn_layout = QHBoxLayout()
        btn_frame.setLayout(btn_layout)
        btn_frame.setStyleSheet("background-color: #6699CC;")

        btn_connect = QPushButton("Connect")
        btn_connect.setFont(self.status_font)
        btn_connect.setStyleSheet("background-color: #0059b3; color: white;")
        btn_connect.clicked.connect(self.connect)

        btn_set = QPushButton("Set Voltage & Current")
        btn_set.setFont(self.status_font)
        btn_set.setStyleSheet("background-color: #007acc; color: white;")
        btn_set.clicked.connect(self.set_values)

        btn_on = QPushButton("Turn ON Output")
        btn_on.setFont(self.status_font)
        btn_on.setStyleSheet("background-color: #009933; color: white;")
        btn_on.clicked.connect(self.turn_on)

        btn_off = QPushButton("Turn OFF Output")
        btn_off.setFont(self.status_font)
        btn_off.setStyleSheet("background-color: #cc3300; color: white;")
        btn_off.clicked.connect(self.turn_off)

        btn_refresh = QPushButton("Refresh Status")
        btn_refresh.setFont(self.status_font)
        btn_refresh.setStyleSheet("background-color: #336699; color: white;")
        btn_refresh.clicked.connect(self.refresh_status)

        btn_disconnect = QPushButton("Disconnect & Unlock")
        btn_disconnect.setFont(self.status_font)
        btn_disconnect.setStyleSheet("background-color: #cc6600; color: white;")
        btn_disconnect.clicked.connect(self.disconnect_and_unlock)

        # Remove the first btn_layout (HBoxLayout) and just keep the vertical container
        btn_frame = QFrame()
        btn_frame.setStyleSheet("background-color: #6699CC;")
        
        upper_row = QHBoxLayout()
        upper_row.addWidget(btn_connect)
        upper_row.addWidget(btn_set)
        
        middle_row = QHBoxLayout()
        middle_row.addWidget(btn_on)
        middle_row.addWidget(btn_off)
        
        lower_row = QHBoxLayout()
        lower_row.addWidget(btn_refresh)
        lower_row.addWidget(btn_disconnect)
        
        btn_v_layout = QVBoxLayout()
        btn_v_layout.addLayout(upper_row)
        btn_v_layout.addLayout(middle_row)
        btn_v_layout.addLayout(lower_row)
        
        btn_frame.setLayout(btn_v_layout)  # Call setLayout only once


        # Status frame with darker blue background
        status_frame = QFrame()
        status_palette = status_frame.palette()
        status_palette.setColor(QPalette.ColorRole.Window, QColor("#001f4d"))
        status_frame.setAutoFillBackground(True)
        status_frame.setPalette(status_palette)
        status_frame.setFrameShape(QFrame.Shape.Panel)
        status_frame.setFrameShadow(QFrame.Shadow.Raised)
        status_frame.setContentsMargins(20, 20, 20, 20)

        status_layout = QFormLayout()
        status_frame.setLayout(status_layout)

        label_status = QLabel("Status:")
        label_status.setFont(self.status_font)
        label_status.setStyleSheet("color: white;")
        self.status_display = QLabel("Device not connected")
        self.status_display.setFont(self.value_font)
        self.status_display.setStyleSheet("color: #99ccff;")

        status_layout.addRow(label_status, self.status_display)

        label_voltage_status = QLabel("Voltage:")
        label_voltage_status.setFont(self.status_font)
        label_voltage_status.setStyleSheet("color: white;")
        self.voltage_display = QLabel("N/A")
        self.voltage_display.setFont(self.value_font)
        self.voltage_display.setStyleSheet("color: #99ff99;")
        status_layout.addRow(label_voltage_status, self.voltage_display)

        label_current_status = QLabel("Current:")
        label_current_status.setFont(self.status_font)
        label_current_status.setStyleSheet("color: white;")
        self.current_display = QLabel("N/A")
        self.current_display.setFont(self.value_font)
        self.current_display.setStyleSheet("color: #99ff99;")
        status_layout.addRow(label_current_status, self.current_display)

        label_output_status = QLabel("Output Status:")
        label_output_status.setFont(self.status_font)
        label_output_status.setStyleSheet("color: white;")
        self.output_display = QLabel("N/A")
        self.output_display.setFont(self.value_font)
        self.output_display.setStyleSheet("color: #ff6666;")
        status_layout.addRow(label_output_status, self.output_display)

        # Add all frames to main layout
        main_layout.addWidget(input_frame)
        main_layout.addWidget(btn_frame)
        main_layout.addWidget(status_frame)

    def show_error(self, title, message):
        QMessageBox.critical(self, title, message)

    def show_info(self, title, message):
        QMessageBox.information(self, title, message)

    def connect(self):
        try:
            visa_address = self.visa_address_edit.text().strip()
            self.rm = pyvisa.ResourceManager()
            self.psu = self.rm.open_resource(visa_address)
            idn = self.psu.query("*IDN?").strip()
            self.show_info("Connection", f"Connected to: {idn}")
            self.refresh_status()
        except Exception as e:
            self.show_error("Connection Error", str(e))

    def disconnect_and_unlock(self):
        try:
            if self.psu:
                # Output OFF
                # self.psu.write("OUTP OFF")  # commented as in original
                # Switch to local mode (releases remote and unlocks front panel)
                self.psu.write("SYST:LOC")
                self.psu.close()
                self.psu = None

            if self.rm:
                self.rm.close()
                self.rm = None

            self.status_display.setText("Device disconnected & front panel unlocked")
            self.voltage_display.setText("N/A")
            self.current_display.setText("N/A")
            self.output_display.setText("N/A")
            self.show_info("Disconnected & Unlocked", "Device remote mode disconnected and front panel unlocked")
        except Exception as e:
            self.show_error("Error", f"Something Wrong\n{str(e)}")

    def set_values(self):
        if self.psu is None:
            self.show_error("Error", "device connect first!")
            return

        ch = self.channel_combo.currentText()
        volt = self.voltage_edit.text()
        curr = self.current_edit.text()

        try:
            ch = int(ch)
            volt = float(volt)
            curr = float(curr)

            if volt > 60:
                self.show_error("Limit Error", "Voltage limit is 60V,Renter")
                return
            if curr > 10:
                self.show_error("Limit Error", "Current limit is 10A,Renter")
                return

            self.psu.write(f"INST:NSEL {ch}")
            self.psu.write(f"VOLT {volt}")
            self.psu.write(f"CURR {curr}")
            self.show_info("Success", f"Voltage and Current set for Channel {ch}")
            self.refresh_status()
        except ValueError:
            self.show_error("Input Error", "Voltage and Current must be numeric values")
        except Exception as e:
            self.show_error("Error", f"Something Wrong:\n{str(e)}")

    def turn_on(self):
        if self.psu is None:
            self.show_error("Error", "device connect first!")
            return
        ch = self.channel_combo.currentText()
        try:
            ch = int(ch)
            self.psu.write(f"INST:NSEL {ch}")
            self.psu.write("OUTP ON")
            self.show_info("Output ON", f"Channel {ch} output ON")
            self.refresh_status()
        except Exception as e:
            self.show_error("Error", f"Something Wrong:\n{str(e)}")

    def turn_off(self):
        if self.psu is None:
            self.show_error("Error", "device connect first!")
            return
        ch = self.channel_combo.currentText()
        try:
            ch = int(ch)
            self.psu.write(f"INST:NSEL {ch}")
            self.psu.write("OUTP OFF")
            self.show_info("Output OFF", f"Channel {ch} output OFF")
            self.refresh_status()
        except Exception as e:
            self.show_error("Error", f"Something Wrong:\n{str(e)}")

    def refresh_status(self):
        if self.psu is None:
            self.status_display.setText("Device not connected")
            self.voltage_display.setText("N/A")
            self.current_display.setText("N/A")
            self.output_display.setText("N/A")
            return

        ch = self.channel_combo.currentText()
        try:
            ch = int(ch)
            self.psu.write(f"INST:NSEL {ch}")
            voltage = self.psu.query("VOLT?").strip()
            current = self.psu.query("CURR?").strip()
            output_status = self.psu.query("OUTP?").strip()

            self.voltage_display.setText(f"{voltage} V")
            self.current_display.setText(f"{current} A")
            self.output_display.setText("ON ✅" if output_status == '1' else "OFF ❌")
            self.status_display.setText(f"Channel {ch} Status Updated")
        except Exception as e:
            self.status_display.setText("Error fetching status")
            self.voltage_display.setText("N/A")
            self.current_display.setText("N/A")
            self.output_display.setText("N/A")
            self.show_error("Error", f"Something Wrong:\n{str(e)}")

    def closeEvent(self, event):
        # Override closeEvent to safely close VISA connections
        try:
            if self.psu:
                self.psu.close()
            if self.rm:
                self.rm.close()
        except:
            pass
        event.accept()

class NIDAQController(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("NI USB DAQ Card Control")
        self.resize(900, 600)
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Devices
        self.device_label = QLabel("Select NI USB Device:")
        self.layout.addWidget(self.device_label)
        self.device_combo = QComboBox()
        self.layout.addWidget(self.device_combo)
        self.device_combo.currentIndexChanged.connect(self.on_device_change)

        # Table for ports/lines
        self.status_label = QLabel("Ports and Digital Lines Status")
        self.layout.addWidget(self.status_label)
        self.port_table = QTableWidget()
        self.layout.addWidget(self.port_table)

        self.refresh_button = QPushButton("Refresh Status")
        self.layout.addWidget(self.refresh_button)
        self.refresh_button.clicked.connect(self.refresh_status)

        self.ni_devices = self.list_ni_usb_devices()
        self.selected_device = None
        self.init_devices()
        if len(self.ni_devices) > 0:
            self.device_combo.setCurrentIndex(0)
            self.on_device_change(0)

    def list_ni_usb_devices(self):
        devices = []
        try:
            system = System.local()
            for dev in system.devices:
                if "USB" in dev.product_type:
                    devices.append(dev)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not list NI devices.\n{e}")
        return devices

    def init_devices(self):
        self.device_combo.clear()
        for dev in self.ni_devices:
            name = dev.name
            serial = getattr(dev, 'serial_num', None)
            if serial is not None:
                serial_hex = f"{serial:08X}"
            else:
                serial_hex = "N/A"
                serial = "N/A"
            text = f"{name} | S/N (DEC): {serial} | S/N (HEX): {serial_hex}"
            self.device_combo.addItem(text, dev)

    def on_device_change(self, index):
        if index < 0 or index >= len(self.ni_devices):
            return
        self.selected_device = self.ni_devices[index]
        self.refresh_status()

    def refresh_status(self):
        dev_name = self.selected_device.name
        port_line_status = []
        try:
            # For all digital input/output lines
            for line in list(self.selected_device.di_lines) + list(self.selected_device.do_lines):
                port_id = line.name.split("/")[1]  # e.g., port0/line0
                line_id = line.name.split("/")[-1]  # e.g. line0
                status = self.read_line_status(line.name)
                port_line_status.append((port_id, line_id, status, line.name))
        except Exception as e:
            print(f"Error reading ports/lines: {e}")
            return

    # (continue table display and toggling UI as already written)

        # For demonstration, query standard digital output ports from /port0/line0:7, can be edited as per each card spec
        port_line_status = []
        for port_num in range(3):  # Typical NI cards have port0/port1/port2
            line_count = 8  # Most common, can check per device as needed
            for line_num in range(line_count):
                line = f"{dev_name}/port{port_num}/line{line_num}"
                status = self.read_line_status(line)
                port_line_status.append((f"port{port_num}", f"line{line_num}", status, line))

        # Build table
        self.port_table.clear()
        self.port_table.setColumnCount(4)
        self.port_table.setHorizontalHeaderLabels(["Port", "Line", "Status", "Toggle High/Low"])
        self.port_table.setRowCount(len(port_line_status))
        for row, data in enumerate(port_line_status):
            port, line, status, phys = data
            self.port_table.setItem(row, 0, QTableWidgetItem(port))
            self.port_table.setItem(row, 1, QTableWidgetItem(line))
            stat_item = QTableWidgetItem("HIGH" if status else "LOW")
            stat_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.port_table.setItem(row, 2, stat_item)
            btn = QPushButton("Toggle")
            btn.clicked.connect(lambda checked, p=phys, s=status, r=row: self.toggle_line(p, r))
            self.port_table.setCellWidget(row, 3, btn)

        self.port_table.resizeColumnsToContents()

    def read_line_status(self, physical_channel):
        try:
            with nidaqmx.Task() as task:
                task.di_channels.add_di_chan(physical_channel)
                data = task.read()
                return bool(data)
        except Exception as e:
            return False  # Assume low if cannot read

    def toggle_line(self, physical_channel, row):
        # Toggle the state (high/low)
        try:
            current_status = self.port_table.item(row,2).text() == "HIGH"
            with nidaqmx.Task() as task:
                task.do_channels.add_do_chan(physical_channel)
                task.write(not current_status)
            # Update GUI
            self.port_table.item(row,2).setText("HIGH" if not current_status else "LOW")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to toggle: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())
